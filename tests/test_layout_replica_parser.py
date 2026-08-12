from dataclasses import FrozenInstanceError
from pathlib import Path
import sys
from types import ModuleType, SimpleNamespace
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


def _make_layout_pdf(path: Path) -> None:
    c = canvas.Canvas(str(path), pagesize=letter)
    c.setFont("Helvetica", 10)
    c.drawString(40, 742, "Example Bank")
    c.drawString(40, 724, "Statement Details")
    c.drawString(40, 690, "Date")
    c.drawString(120, 690, "Description")
    c.drawString(330, 690, "Debit")
    c.drawString(410, 690, "Credit")
    c.drawString(500, 690, "Balance")
    c.drawString(40, 670, "01/01/2026")
    c.drawString(120, 670, "OPENING BALANCE")
    c.drawString(500, 670, "1,234.50 CR")
    c.drawString(40, 650, "02/01/2026")
    c.drawString(120, 650, "CARD PAYMENT ABC-123")
    c.drawString(330, 650, "25.75")
    c.drawString(500, 650, "1,208.75 CR")
    c.save()


def _make_layout_line(page_num, index, words, source="pdf-text"):
    from parsers.layout_replica_parser import LayoutLine, LayoutWord

    layout_words = [
        LayoutWord(
            text=text,
            x0=x0,
            x1=x1,
            top=index * 10.0,
            bottom=index * 10.0 + 8.0,
            page=page_num,
            source=source,
        )
        for text, x0, x1 in words
    ]
    return LayoutLine(
        page=page_num,
        index=index,
        top=index * 10.0,
        bottom=index * 10.0 + 8.0,
        center_y=index * 10.0 + 4.0,
        words=layout_words,
        text=" ".join(word.text for word in layout_words),
    )


def _make_layout_page(page_num, lines, source="pdf-text"):
    from parsers.layout_replica_parser import LayoutPage

    return LayoutPage(
        page_number=page_num,
        width=620,
        height=800,
        source=source,
        words=[word for line in lines for word in line.words],
        lines=lines,
    )


def test_layout_replica_parser_preserves_visible_lines_and_strings(tmp_path):
    from parsers.layout_replica_parser import LayoutReplicaParser

    pdf_path = tmp_path / "statement.pdf"
    _make_layout_pdf(pdf_path)

    parser = LayoutReplicaParser(use_ocr=False)
    parser.parse(str(pdf_path), "statement.pdf")

    assert parser.extraction_metadata.has_data is True
    assert parser.extraction_metadata.extraction_method == "layout_replica"
    assert parser.extraction_metadata.pdf_type == "text"
    assert parser.extraction_metadata.confidence == "good"
    quality_report = parser.get_quality_report()
    assert quality_report["source_line_count"] == sum(len(page.lines) for page in parser.pages)
    assert quality_report["source_word_count"] == sum(len(page.words) for page in parser.pages)
    assert quality_report["source_coverage_pct"] == 100.0
    assert quality_report["accuracy_proxy_pct"] == 0.0
    assert quality_report["review_required"] is False
    assert parser.raw_table is not None
    assert [column.header for column in parser.table_columns] == [
        "Date",
        "Description",
        "Debit",
        "Credit",
        "Balance",
    ]
    assert parser.table_rows[1].values == [
        "02/01/2026",
        "CARD PAYMENT ABC-123",
        "25.75",
        "",
        "1,208.75 CR",
    ]

    raw_lines = [" ".join(row) for row in parser.raw_table["rows"]]
    assert any("Statement Details" in line for line in raw_lines)
    assert any("01/01/2026" in line and "OPENING BALANCE" in line for line in raw_lines)

    output_path = tmp_path / "replica.xlsx"
    parser.write_excel(str(output_path))

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Exact_Copy", "Table_Data", "Full_Text"]

    table_sheet = wb["Table_Data"]
    assert [table_sheet.cell(1, col).value for col in range(1, 6)] == [
        "Date",
        "Description",
        "Debit",
        "Credit",
        "Balance",
    ]
    assert [table_sheet.cell(3, col).value for col in range(1, 6)] == [
        "02/01/2026",
        "CARD PAYMENT ABC-123",
        "25.75",
        None,
        "1,208.75 CR",
    ]

    table_values = [
        cell.value
        for row in table_sheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "1,234.50 CR" in table_values
    assert "25.75" in table_values

    amount_cell = next(cell for row in table_sheet.iter_rows() for cell in row if cell.value == "25.75")
    assert amount_cell.number_format == "@"


def test_layout_replica_workbook_exports_table_and_full_text_sheets(tmp_path):
    from parsers.layout_replica_parser import LayoutReplicaParser

    pdf_path = tmp_path / "statement.pdf"
    _make_layout_pdf(pdf_path)

    parser = LayoutReplicaParser(use_ocr=False)
    parser.parse(str(pdf_path), "statement.pdf")
    output_path = tmp_path / "replica.xlsx"
    parser.write_excel(str(output_path))

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Exact_Copy", "Table_Data", "Full_Text"]
    sheet = wb["Table_Data"]
    assert sheet.max_column == 5
    assert sheet.max_row == 3

    full_text = wb["Full_Text"]
    assert [full_text.cell(1, col).value for col in range(1, 5)] == [
        "Page", "Line", "Source", "Text",
    ]
    total_lines = sum(len(page.lines) for page in parser.pages)
    assert full_text.max_row == total_lines + 1  # header + every visual line

    texts = [full_text.cell(row, 4).value for row in range(2, full_text.max_row + 1)]
    # content the table sheet drops must still be in the workbook
    assert any("Example Bank" in (t or "") for t in texts)
    assert any("Statement Details" in (t or "") for t in texts)
    # table content appears too (reading order, untouched)
    assert any("CARD PAYMENT ABC-123" in (t or "") for t in texts)


def test_exact_copy_is_first_and_covers_every_line_and_word_once(tmp_path):
    from parsers.layout_replica_parser import LayoutReplicaParser

    page_one_lines = [
        _make_layout_line(1, 1, [
            ("P1L1-A", 12.0, 28.0),
            ("P1L1-B", 13.0, 29.0),  # same grid cell; neither token may be lost
            ("P1L1-C", 120.0, 145.0),
        ]),
        _make_layout_line(1, 2, [("P1L2-A", 48.0, 70.0)]),
    ]
    page_two_lines = [
        _make_layout_line(2, 1, [
            ("P2L1-A", 18.0, 38.0),
            ("P2L1-B", 180.0, 205.0),
        ]),
    ]
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [
        _make_layout_page(1, page_one_lines),
        _make_layout_page(2, page_two_lines),
    ]

    output_path = tmp_path / "exact-copy.xlsx"
    parser.write_excel(str(output_path))
    wb = load_workbook(output_path)

    assert wb.sheetnames == ["Exact_Copy", "Table_Data", "Full_Text"]
    exact = wb["Exact_Copy"]
    assert exact.max_row == 5  # two page separators plus three source lines
    assert exact.cell(1, 1).value == "Page 1"
    assert exact.cell(4, 1).value == "Page 2"

    source_lines = page_one_lines + page_two_lines
    exact_rows = [2, 3, 5]
    for worksheet_row, source_line in zip(exact_rows, source_lines):
        written_tokens = []
        for cell in exact[worksheet_row]:
            if cell.value:
                written_tokens.extend(str(cell.value).split())
        assert written_tokens == [word.text for word in source_line.words]

    expected_tokens = [word.text for line in source_lines for word in line.words]
    written_tokens = []
    for worksheet_row in exact_rows:
        for cell in exact[worksheet_row]:
            if cell.value:
                written_tokens.extend(str(cell.value).split())
    assert sorted(written_tokens) == sorted(expected_tokens)
    assert len(written_tokens) == len(expected_tokens)

    # Existing page-grid mapping remains authoritative for horizontal placement.
    first_word = page_one_lines[0].words[0]
    mapped_column = parser._column_for_x(first_word.x0, parser.pages[0].width)
    assert exact.cell(2, mapped_column).value == "P1L1-A P1L1-B"

    full_text = wb["Full_Text"]
    assert [
        (full_text.cell(row, 1).value, full_text.cell(row, 2).value, full_text.cell(row, 4).value)
        for row in range(2, full_text.max_row + 1)
    ] == [
        ("1", "1", "P1L1-A P1L1-B P1L1-C"),
        ("1", "2", "P1L2-A"),
        ("2", "1", "P2L1-A P2L1-B"),
    ]


def test_source_strings_beginning_with_equals_are_never_excel_formulas(tmp_path):
    from parsers.layout_replica_parser import (
        LayoutReplicaParser,
        TableColumn,
        TableReplicaRow,
    )

    line = _make_layout_line(1, 1, [
        ("=2+2", 0.0, 24.0),
        ("note", 90.0, 112.0),
    ])
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [_make_layout_page(1, [line])]
    parser.table_columns = [
        TableColumn("=Date", 0.0, 30.0, 0.0, 60.0),
        TableColumn("Description", 90.0, 150.0, 60.0, 180.0),
    ]
    parser.table_rows = [
        TableReplicaRow(
            page=1,
            line=1,
            values=["=2+2", '=HYPERLINK("https://example.invalid")'],
            source="pdf-text",
            schema_headers=("=Date", "Description"),
            schema_signature=("=date", "description"),
        )
    ]

    output_path = tmp_path / "formula-safe.xlsx"
    parser.write_excel(str(output_path))
    wb = load_workbook(output_path, data_only=False)

    expected_formula_like_text = {
        "Exact_Copy": {"=2+2"},
        "Table_Data": {
            "=Date",
            "=2+2",
            '=HYPERLINK("https://example.invalid")',
        },
        "Full_Text": {"=2+2 note"},
    }
    for sheet_name, expected_values in expected_formula_like_text.items():
        cells = [
            cell
            for row in wb[sheet_name].iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        assert {cell.value for cell in cells} == expected_values
        assert all(cell.data_type == "s" for cell in cells)

    with ZipFile(output_path) as archive:
        worksheet_xml = "".join(
            archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
    assert "<f" not in worksheet_xml


def test_schema_changes_export_separately_without_truncating_earlier_rows(tmp_path):
    from parsers.layout_replica_parser import LayoutReplicaParser

    # Page 1 has no header, so its seven-column transaction layout is inferred.
    first_page = _make_layout_page(1, [
        _make_layout_line(1, 1, [
            ("29/01/25", 20.0, 65.0),
            ("PAYMENT", 90.0, 145.0),
            ("REF-991", 190.0, 235.0),
            ("29/01/25", 285.0, 330.0),
            ("72.00", 380.0, 415.0),
            ("0.00", 460.0, 490.0),
            ("1,523.00", 545.0, 595.0),
        ]),
    ])
    # Page 2 introduces a real, narrower five-column schema. Historically this
    # became the workbook-wide schema and sliced the two rightmost page-1 cells.
    second_page = _make_layout_page(2, [
        _make_layout_line(2, 1, [
            ("Date", 20.0, 45.0),
            ("Description", 100.0, 165.0),
            ("Debit", 330.0, 360.0),
            ("Credit", 410.0, 445.0),
            ("Balance", 510.0, 555.0),
        ]),
        _make_layout_line(2, 2, [
            ("30/01/25", 20.0, 65.0),
            ("CARD-PAYMENT", 100.0, 180.0),
            ("25.75", 330.0, 360.0),
            ("0.00", 410.0, 440.0),
            ("1,497.25", 510.0, 560.0),
        ]),
    ])

    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [first_page, second_page]
    parser._build_table_replica()

    assert len(parser.table_rows) == 2
    first_row, second_row = parser.table_rows
    assert first_row.schema_headers == tuple(f"Column {idx}" for idx in range(1, 8))
    assert second_row.schema_headers == ("Date", "Description", "Debit", "Credit", "Balance")
    assert isinstance(first_row.schema_headers, tuple)
    assert isinstance(first_row.schema_signature, tuple)
    assert first_row.schema_signature != second_row.schema_signature
    with pytest.raises(FrozenInstanceError):
        first_row.schema_headers = ("Changed",)

    parser._build_metadata()
    quality_report = parser._build_quality_report()
    assert parser.extraction_metadata.confidence == "low"
    assert quality_report["table_schema_count"] == 2
    assert quality_report["review_required"] is True
    assert quality_report["source_coverage_pct"] == 100.0
    assert quality_report["accuracy_proxy_pct"] == 0.0

    output_path = tmp_path / "schema-groups.xlsx"
    parser.write_excel(str(output_path))
    wb = load_workbook(output_path)

    assert wb.sheetnames == ["Exact_Copy", "Table_Data", "Table_Data_2", "Full_Text"]
    first_table = wb["Table_Data"]
    assert [first_table.cell(1, col).value for col in range(1, 8)] == [
        f"Column {idx}" for idx in range(1, 8)
    ]
    assert [first_table.cell(2, col).value for col in range(1, 8)] == first_row.values
    assert first_table.max_column == 7

    second_table = wb["Table_Data_2"]
    assert [second_table.cell(1, col).value for col in range(1, 6)] == [
        "Date", "Description", "Debit", "Credit", "Balance",
    ]
    assert [second_table.cell(2, col).value for col in range(1, 6)] == second_row.values
    assert second_table.max_column == 5


def test_table_replica_keeps_rows_before_repeated_header_on_same_pdf_page():
    from parsers.layout_replica_parser import LayoutLine, LayoutPage, LayoutReplicaParser, LayoutWord

    def make_line(page_num, index, words):
        layout_words = [
            LayoutWord(
                text=text,
                x0=x0,
                x1=x1,
                top=index * 10.0,
                bottom=index * 10.0 + 8.0,
                page=page_num,
                source="pdf-text",
            )
            for text, x0, x1 in words
        ]
        return LayoutLine(
            page=page_num,
            index=index,
            top=index * 10.0,
            bottom=index * 10.0 + 8.0,
            center_y=index * 10.0 + 4.0,
            words=layout_words,
            text=" ".join(word.text for word in layout_words),
        )

    def make_page(page_num, lines):
        return LayoutPage(
            page_number=page_num,
            width=620,
            height=800,
            source="pdf-text",
            words=[word for line in lines for word in line.words],
            lines=lines,
        )

    separator = [("-" * 100, 16.6, 597.2)]
    header = [
        ("TXN", 16.6, 32.4),
        ("DT", 37.7, 48.2),
        ("VALUE_DT", 64.1, 106.3),
        ("BRN", 116.9, 132.7),
        ("DESCRIPTION", 169.6, 227.7),
        ("REFERENCE", 280.5, 328.0),
        ("DEBITS", 380.8, 412.5),
        ("CREDITS", 454.7, 491.6),
        ("BALANCE", 560.3, 597.2),
    ]
    first_page_lines = [
        make_line(1, 1, separator),
        make_line(1, 2, header),
        make_line(1, 3, separator),
        make_line(1, 4, [
            ("01/04/19", 16.6, 58.8),
            ("01/04/19", 64.1, 106.3),
            ("1763", 111.6, 132.7),
            ("OPENING", 143.2, 183.0),
            ("BALANCE", 187.0, 227.0),
            ("1,00,000.00", 533.9, 591.9),
        ]),
    ]
    second_page_lines = [
        make_line(2, 1, [
            ("13/04/19", 16.6, 58.8),
            ("13/04/19", 64.1, 106.3),
            ("1763", 111.6, 132.7),
            ("MPAY/UPI/FI", 143.2, 206.0),
            ("Funds", 210.0, 240.0),
            ("Trans-1", 244.0, 275.2),
            ("772108174541", 280.5, 343.8),
            ("8,600.00", 375.5, 423.0),
            ("42,557.10", 533.9, 591.9),
        ]),
        make_line(2, 2, [("684155000075893", 143.2, 217.0)]),
        make_line(2, 3, separator),
        make_line(2, 4, [("page", 280.0, 302.0), (":", 306.0, 310.0), ("2", 316.0, 322.0)]),
        make_line(2, 5, [
            ("STATEMENT", 170.0, 220.0),
            ("OF", 225.0, 238.0),
            ("ACCOUNT", 243.0, 285.0),
        ]),
        make_line(2, 6, [
            ("Account", 280.0, 315.0),
            ("Number", 320.0, 355.0),
            ("1684135000012107", 360.0, 450.0),
        ]),
        make_line(2, 7, separator),
        make_line(2, 8, header),
        make_line(2, 9, separator),
        make_line(2, 10, [
            ("15/04/19", 16.6, 58.8),
            ("15/04/19", 64.1, 106.3),
            ("1763", 111.6, 132.7),
            ("ATM", 143.2, 162.0),
            ("Withdrawal", 166.0, 222.0),
            ("123456789012", 280.5, 343.8),
            ("2,000.00", 375.5, 423.0),
            ("40,557.10", 533.9, 591.9),
        ]),
    ]

    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [
        make_page(1, first_page_lines),
        make_page(2, second_page_lines),
    ]
    parser._build_table_replica()

    page_two_rows = [
        row.values for row in parser.table_rows
        if row.page == 2
    ]
    assert [
        "13/04/19",
        "13/04/19",
        "1763",
        "MPAY/UPI/FI Funds Trans-1 684155000075893",
        "772108174541",
        "8,600.00",
        "",
        "42,557.10",
    ] in page_two_rows
    assert not any("STATEMENT OF ACCOUNT" in " ".join(row) for row in page_two_rows)
    assert not any("Account Number" in " ".join(row) for row in page_two_rows)
    # continuation line must no longer be its own row
    assert not any(row[0] == "" and "684155000075893" in " ".join(row) for row in page_two_rows)


def test_table_replica_infers_columns_when_header_is_missing():
    from parsers.layout_replica_parser import LayoutLine, LayoutPage, LayoutReplicaParser, LayoutWord

    def make_line(index, words):
        layout_words = [
            LayoutWord(
                text=text,
                x0=x0,
                x1=x1,
                top=index * 10.0,
                bottom=index * 10.0 + 8.0,
                page=1,
                source="ocr",
            )
            for text, x0, x1 in words
        ]
        return LayoutLine(
            page=1,
            index=index,
            top=index * 10.0,
            bottom=index * 10.0 + 8.0,
            center_y=index * 10.0 + 4.0,
            words=layout_words,
            text=" ".join(word.text for word in layout_words),
        )

    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [
        LayoutPage(
            page_number=1,
            width=600,
            height=800,
            source="ocr",
            words=[],
            lines=[
                make_line(1, [("From", 5, 25), (":", 28, 32), ("01/04/2018", 40, 90)]),
                make_line(2, [
                    ("29/01/25", 6, 41),
                    ("UPI-SELVI", 47, 176),
                    ("0000024036557328", 266, 332),
                    ("29/01/25", 338, 370),
                    ("72.00", 426, 449),
                    ("0.00", 508, 526),
                    ("1,523.00", 570, 596),
                ]),
                make_line(3, [("PTY-YESB0MCHUPI-024036557328-PAYMENT", 47, 218)]),
                make_line(4, [("Closing balance includes funds earmarked for hold", 47, 300)]),
            ],
        )
    ]

    parser._build_table_replica()

    assert [column.header for column in parser.table_columns] == [
        "Column 1",
        "Column 2",
        "Column 3",
        "Column 4",
        "Column 5",
        "Column 6",
        "Column 7",
    ]
    assert parser.table_rows[0].values == [
        "29/01/25",
        "UPI-SELVI PTY-YESB0MCHUPI-024036557328-PAYMENT",
        "0000024036557328",
        "29/01/25",
        "72.00",
        "0.00",
        "1,523.00",
    ]
    assert len(parser.table_rows) == 1


def test_continuation_merges_into_description_and_reference_columns():
    from parsers.layout_replica_parser import LayoutLine, LayoutPage, LayoutReplicaParser, LayoutWord

    def make_line(page_num, index, words):
        layout_words = [
            LayoutWord(text=text, x0=x0, x1=x1, top=index * 10.0,
                       bottom=index * 10.0 + 8.0, page=page_num, source="pdf-text")
            for text, x0, x1 in words
        ]
        return LayoutLine(page=page_num, index=index, top=index * 10.0,
                          bottom=index * 10.0 + 8.0, center_y=index * 10.0 + 4.0,
                          words=layout_words,
                          text=" ".join(word.text for word in layout_words))

    separator = [("-" * 100, 16.6, 597.2)]
    header = [
        ("TXN", 16.6, 32.4), ("DT", 37.7, 48.2), ("VALUE_DT", 64.1, 106.3),
        ("BRN", 116.9, 132.7), ("DESCRIPTION", 169.6, 227.7),
        ("REFERENCE", 280.5, 328.0), ("DEBITS", 380.8, 412.5),
        ("CREDITS", 454.7, 491.6), ("BALANCE", 560.3, 597.2),
    ]
    lines = [
        make_line(1, 1, separator),
        make_line(1, 2, header),
        make_line(1, 3, separator),
        make_line(1, 4, [
            ("03/04/19", 16.6, 58.8), ("03/04/19", 64.1, 106.3), ("1763", 111.6, 132.7),
            ("IMPS", 143.2, 168.0), ("DR-1763308", 172.0, 227.0),
            ("909223247145", 280.5, 343.8), ("5,000.00", 375.5, 423.0),
            ("1,66,685.70", 533.9, 591.9),
        ]),
        # wraps into DESCRIPTION and REFERENCE columns, no date, no amounts
        make_line(1, 5, [("HDFC0000240-3017FA", 143.2, 230.0), ("835", 280.5, 300.0)]),
    ]
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [LayoutPage(page_number=1, width=620, height=800, source="pdf-text",
                               words=[w for line in lines for w in line.words], lines=lines)]
    parser._build_table_replica()

    assert len(parser.table_rows) == 1
    # 8 columns: "TXN" and "DT" merge into one header cell, so
    # idx 3 = DESCRIPTION, idx 4 = REFERENCE, idx 5 = DEBITS
    row = parser.table_rows[0].values
    assert row[3] == "IMPS DR-1763308 HDFC0000240-3017FA"
    assert row[4] == "909223247145 835"
    assert row[5] == "5,000.00"


def test_date_starting_line_is_never_merged_as_continuation():
    from parsers.layout_replica_parser import LayoutLine, LayoutPage, LayoutReplicaParser, LayoutWord

    def make_line(page_num, index, words):
        layout_words = [
            LayoutWord(text=text, x0=x0, x1=x1, top=index * 10.0,
                       bottom=index * 10.0 + 8.0, page=page_num, source="pdf-text")
            for text, x0, x1 in words
        ]
        return LayoutLine(page=page_num, index=index, top=index * 10.0,
                          bottom=index * 10.0 + 8.0, center_y=index * 10.0 + 4.0,
                          words=layout_words,
                          text=" ".join(word.text for word in layout_words))

    separator = [("-" * 100, 16.6, 597.2)]
    header = [
        ("TXN", 16.6, 32.4), ("DT", 37.7, 48.2), ("VALUE_DT", 64.1, 106.3),
        ("BRN", 116.9, 132.7), ("DESCRIPTION", 169.6, 227.7),
        ("REFERENCE", 280.5, 328.0), ("DEBITS", 380.8, 412.5),
        ("CREDITS", 454.7, 491.6), ("BALANCE", 560.3, 597.2),
    ]
    lines = [
        make_line(1, 1, separator),
        make_line(1, 2, header),
        make_line(1, 3, separator),
        make_line(1, 4, [
            ("01/04/19", 16.6, 58.8), ("01/04/19", 64.1, 106.3), ("1763", 111.6, 132.7),
            ("ATM", 143.2, 162.0), ("CSW", 166.0, 190.0),
            ("909110624882", 280.5, 343.8), ("10,000.00", 375.5, 423.0),
            ("1,43,265.70", 533.9, 591.9),
        ]),
        make_line(1, 5, [
            ("01/04/19", 16.6, 58.8), ("01/04/19", 64.1, 106.3), ("1763", 111.6, 132.7),
            ("CA", 143.2, 155.0), ("ATM", 159.0, 178.0), ("TXN", 182.0, 200.0),
            ("909110624882", 280.5, 343.8), ("20.00", 375.5, 423.0),
            ("1,43,245.70", 533.9, 591.9),
        ]),
    ]
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [LayoutPage(page_number=1, width=620, height=800, source="pdf-text",
                               words=[w for line in lines for w in line.words], lines=lines)]
    parser._build_table_replica()

    assert len(parser.table_rows) == 2
    assert parser.table_rows[0].values[5] == "10,000.00"
    assert parser.table_rows[1].values[5] == "20.00"


def test_multi_token_ocr_boxes_are_split_per_token():
    from parsers.layout_replica_parser import LayoutReplicaParser

    parser = LayoutReplicaParser(use_ocr=False)
    words = parser._ocr_results_to_layout_words(
        [
            # one Paddle det box spanning three table cells
            {"text": "03/09/18 03/09/18 1763", "bbox": [100.0, 10.0, 330.0, 22.0], "confidence": 0.9},
            {"text": "single", "bbox": [400.0, 10.0, 440.0, 22.0], "confidence": 0.8},
        ],
        page_num=1,
        page_width=612.0,
        page_height=792.0,
        image_width=612,
        image_height=792,
        source="ocr",
    )

    assert [w.text for w in words] == ["03/09/18", "03/09/18", "1763", "single"]
    d1, d2, brn, single = words
    # tokens keep left-to-right order inside the original box
    assert 100.0 == d1.x0 and d1.x1 < d2.x0 and d2.x1 < brn.x0
    assert abs(brn.x1 - 330.0) < 1e-6
    # widths proportional to character counts (8/8/4 chars + 2 gaps)
    assert abs((d1.x1 - d1.x0) - (d2.x1 - d2.x0)) < 1e-6
    assert (d1.x1 - d1.x0) > (brn.x1 - brn.x0)
    # untouched single-token box
    assert (single.x0, single.x1) == (400.0, 440.0)
    assert all(w.confidence == 0.9 for w in (d1, d2, brn))


def test_layout_replica_factory_does_not_initialize_paddle_when_disabled(monkeypatch):
    from parsers.layout_replica_parser import LayoutWord, create_layout_replica_parser

    paddle_initializations = []

    class UnexpectedPaddleProcessor:
        def __init__(self, **kwargs):
            paddle_initializations.append(kwargs)

    fake_paddle_module = ModuleType("parsers.paddleocr_processor")
    fake_paddle_module.PaddleOCRProcessor = UnexpectedPaddleProcessor
    monkeypatch.setitem(sys.modules, "parsers.paddleocr_processor", fake_paddle_module)

    parser = create_layout_replica_parser(use_ocr=True, use_paddleocr=False)
    image = SimpleNamespace(mode="RGB", width=100, height=100)
    fallback_word = LayoutWord(
        text="fallback",
        x0=1,
        x1=20,
        top=1,
        bottom=10,
        page=1,
        source="ocr-tesseract",
        confidence=0.5,
    )
    monkeypatch.setattr(parser, "_render_page_image", lambda *_args: image)
    monkeypatch.setattr(parser, "_extract_tesseract_words", lambda *_args: [fallback_word])

    words = parser._extract_ocr_words("statement.pdf", 1, 100, 100)

    assert parser.use_paddleocr is False
    assert paddle_initializations == []
    assert words == [fallback_word]


def test_paddle_coordinate_extraction_keeps_low_confidence_words(monkeypatch):
    from parsers.layout_replica_parser import LayoutReplicaParser

    thresholds = []

    class FakePaddleProcessor:
        def __init__(self, **_kwargs):
            pass

        def extract_with_coordinates(self, _image, confidence_threshold):
            thresholds.append(confidence_threshold)
            return [
                {
                    "text": "uncertain",
                    "confidence": 0.01,
                    "bbox": [1, 2, 30, 12],
                },
                {"text": "Description", "confidence": 0.9, "bbox": [35, 2, 65, 12]},
                {"text": "Withdrawal", "confidence": 0.9, "bbox": [70, 2, 90, 12]},
                {"text": "Balance", "confidence": 0.9, "bbox": [92, 2, 100, 12]},
            ]

    fake_paddle_module = ModuleType("parsers.paddleocr_processor")
    fake_paddle_module.PaddleOCRProcessor = FakePaddleProcessor
    monkeypatch.setitem(sys.modules, "parsers.paddleocr_processor", fake_paddle_module)

    parser = LayoutReplicaParser(use_ocr=True, use_paddleocr=True)
    image = SimpleNamespace(mode="RGB", width=100, height=100)
    monkeypatch.setattr(parser, "_render_page_image", lambda *_args: image)
    monkeypatch.setattr(
        parser,
        "_extract_tesseract_words",
        lambda *_args: pytest.fail("Tesseract should not run after meaningful Paddle coverage"),
    )

    words = parser._extract_ocr_words("statement.pdf", 1, 100, 100)

    assert thresholds == [0.0]
    assert [word.text for word in words] == [
        "uncertain",
        "Description",
        "Withdrawal",
        "Balance",
    ]
    assert words[0].confidence == 0.01


def test_sparse_paddle_result_is_supplemented_by_tesseract_without_visual_duplicates(
    monkeypatch,
):
    from parsers.layout_replica_parser import LayoutReplicaParser, LayoutWord

    class SparsePaddleProcessor:
        def __init__(self, **_kwargs):
            pass

        def extract_with_coordinates(self, _image, confidence_threshold):
            assert confidence_threshold == 0.0
            return [{"text": "Date", "confidence": 0.2, "bbox": [1, 2, 20, 12]}]

    fake_paddle_module = ModuleType("parsers.paddleocr_processor")
    fake_paddle_module.PaddleOCRProcessor = SparsePaddleProcessor
    monkeypatch.setitem(sys.modules, "parsers.paddleocr_processor", fake_paddle_module)

    parser = LayoutReplicaParser(use_ocr=True, use_paddleocr=True)
    image = SimpleNamespace(mode="RGB", width=100, height=100)
    monkeypatch.setattr(parser, "_render_page_image", lambda *_args: image)
    tesseract_calls = []

    def tesseract_words(*_args):
        tesseract_calls.append(True)
        return [
            # Same visible box, but a different recognition. The primary
            # Paddle token must remain and this must not create a duplicate.
            LayoutWord("Dale", 1, 20, 2, 12, 1, "ocr-tesseract", 0.8),
            # Geometry alone cannot suppress unrelated OCR text: sparse PDF
            # layers sometimes contain stray or hidden overlapping tokens.
            LayoutWord("Balance", 1, 20, 2, 12, 1, "ocr-tesseract", 0.8),
            LayoutWord("Description", 25, 60, 2, 12, 1, "ocr-tesseract", 0.9),
            LayoutWord("Amount", 70, 95, 2, 12, 1, "ocr-tesseract", 0.7),
        ]

    monkeypatch.setattr(parser, "_extract_tesseract_words", tesseract_words)

    words = parser._extract_ocr_words("statement.pdf", 1, 100, 100)

    assert tesseract_calls == [True]
    assert [word.text for word in words] == ["Date", "Balance", "Description", "Amount"]


def test_sparse_native_text_layer_is_supplemented_with_ocr_without_duplicates(monkeypatch):
    import parsers.layout_replica_parser as layout_module
    from parsers.layout_replica_parser import LayoutReplicaParser, LayoutWord

    class FakePdf:
        pages = [
            SimpleNamespace(width=100, height=100),
            SimpleNamespace(width=100, height=100),
        ]

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    fake_pdfplumber = ModuleType("pdfplumber")
    fake_pdfplumber.open = lambda _path: FakePdf()
    monkeypatch.setitem(sys.modules, "pdfplumber", fake_pdfplumber)
    monkeypatch.setattr(layout_module, "raise_if_password_protected", lambda _path: None)

    parser = LayoutReplicaParser(use_ocr=True, use_paddleocr=False)
    monkeypatch.setattr(parser, "validate_pdf_file", lambda _path: None)
    sparse_native = LayoutWord("Date", 1, 20, 2, 12, 1, "pdf-text")
    normal_native = [
        LayoutWord(f"native-{index}", index * 10, index * 10 + 8, 2, 12, 2, "pdf-text")
        for index in range(8)
    ]
    monkeypatch.setattr(
        parser,
        "_extract_pdf_words",
        lambda _page, page_num: [sparse_native] if page_num == 1 else normal_native,
    )
    ocr_calls = []

    def extract_ocr(_path, page_num, _width, _height):
        ocr_calls.append(page_num)
        return [
            # Geometry identifies this as the same visible word despite OCR's
            # misspelling, so the native token remains authoritative.
            LayoutWord("Dale", 1, 20, 2, 12, 1, "ocr-tesseract", 0.7),
            LayoutWord("Description", 25, 60, 2, 12, 1, "ocr-tesseract", 0.9),
            LayoutWord("Amount", 70, 95, 2, 12, 1, "ocr-tesseract", 0.8),
        ]

    monkeypatch.setattr(parser, "_extract_ocr_words", extract_ocr)

    parser.parse("statement.pdf", "statement.pdf")

    assert ocr_calls == [1]
    assert parser.pages[0].source == "hybrid"
    assert [word.text for word in parser.pages[0].words] == [
        "Date",
        "Description",
        "Amount",
    ]
    assert parser.pages[1].source == "pdf-text"
    assert parser.extraction_metadata.pdf_type == "hybrid"
    assert parser.get_quality_report()["ocr_page_count"] == 1
    assert parser.get_quality_report()["text_page_count"] == 2


def test_tesseract_checks_every_psm_and_prefers_word_coverage_without_dropping_low_confidence(
    monkeypatch,
):
    from parsers.layout_replica_parser import LayoutReplicaParser

    def ocr_data(texts, confidences):
        count = len(texts)
        return {
            "text": texts,
            "conf": confidences,
            "left": list(range(0, count * 10, 10)),
            "top": [0] * count,
            "width": [8] * count,
            "height": [8] * count,
        }

    calls = []

    def image_to_data(_image, output_type, config):
        calls.append((output_type, config))
        if "--psm 6" in config:
            return ocr_data(["first-only"], [99])
        return ocr_data(["second", "uncertain", "not-a-word"], [75, 0, -1])

    fake_tesseract = ModuleType("pytesseract")
    fake_tesseract.Output = SimpleNamespace(DICT="dict")
    fake_tesseract.image_to_data = image_to_data
    monkeypatch.setitem(sys.modules, "pytesseract", fake_tesseract)

    parser = LayoutReplicaParser(use_paddleocr=False)
    image = SimpleNamespace(mode="RGB", width=100, height=100)
    words = parser._extract_tesseract_words(image, 1, 100, 100)

    assert [config for _output_type, config in calls] == [
        "--psm 6 -c preserve_interword_spaces=1",
        "--psm 11",
    ]
    assert [word.text for word in words] == ["second", "uncertain"]
    assert words[1].confidence == 0


def test_tesseract_breaks_equal_coverage_ties_by_confidence(monkeypatch):
    from parsers.layout_replica_parser import LayoutReplicaParser

    def image_to_data(_image, output_type, config):
        del output_type
        if "--psm 6" in config:
            texts, confidences = ["low-a", "low-b"], [5, 5]
        else:
            texts, confidences = ["best-a", "best-b"], [80, 0]
        return {
            "text": texts,
            "conf": confidences,
            "left": [0, 10],
            "top": [0, 0],
            "width": [8, 8],
            "height": [8, 8],
        }

    fake_tesseract = ModuleType("pytesseract")
    fake_tesseract.Output = SimpleNamespace(DICT="dict")
    fake_tesseract.image_to_data = image_to_data
    monkeypatch.setitem(sys.modules, "pytesseract", fake_tesseract)

    parser = LayoutReplicaParser(use_paddleocr=False)
    image = SimpleNamespace(mode="RGB", width=100, height=100)
    words = parser._extract_tesseract_words(image, 1, 100, 100)

    assert [word.text for word in words] == ["best-a", "best-b"]


def test_tesseract_preprocessing_preserves_canvas_and_produces_binary_rgb():
    from PIL import Image

    from parsers.layout_replica_parser import LayoutReplicaParser

    image = Image.new("RGB", (12, 8), "white")
    for x, y in [(1, 1), (2, 5), (5, 3), (9, 6), (10, 2)]:
        image.putpixel((x, y), (0, 0, 0))

    processed = LayoutReplicaParser._preprocess_tesseract_image(image)

    assert processed is not None
    assert processed.size == image.size
    assert processed.mode == "RGB"
    assert {
        channel
        for pixel in processed.getdata()
        for channel in pixel
    } <= {0, 255}


def test_tesseract_preprocessed_candidate_wins_on_table_coverage_not_junk_word_count(
    monkeypatch,
):
    from parsers.layout_replica_parser import LayoutReplicaParser

    original_image = SimpleNamespace(mode="RGB", width=200, height=200, variant="original")
    processed_image = SimpleNamespace(mode="RGB", width=200, height=200, variant="processed")

    def to_data(entries):
        return {
            "text": [entry[0] for entry in entries],
            "conf": [80] * len(entries),
            "left": [entry[1] for entry in entries],
            "top": [entry[2] for entry in entries],
            "width": [8] * len(entries),
            "height": [8] * len(entries),
        }

    junk_psm_6 = [(f"noise-{index}", index % 20 * 9, index // 20 * 12) for index in range(40)]
    junk_psm_11 = [(f"artifact-{index}", index % 20 * 9, index // 20 * 12) for index in range(50)]
    table_entries = [
        ("Date", 2, 0),
        ("Description", 35, 0),
        ("Debit", 125, 0),
        ("Balance", 165, 0),
    ]
    for row in range(1, 9):
        top = row * 15
        table_entries.extend([
            (f"2026-01-{row:02d}", 2, top),
            ("PAYMENT", 35, top),
            (f"{row}.00", 125, top),
            (f"{100 - row}.00", 165, top),
        ])

    calls = []

    def image_to_data(image, output_type, config):
        del output_type
        calls.append((image.variant, config))
        if image.variant == "processed":
            return to_data(table_entries)
        if "--psm 6" in config:
            return to_data(junk_psm_6)
        return to_data(junk_psm_11)

    fake_tesseract = ModuleType("pytesseract")
    fake_tesseract.Output = SimpleNamespace(DICT="dict")
    fake_tesseract.image_to_data = image_to_data
    monkeypatch.setitem(sys.modules, "pytesseract", fake_tesseract)

    parser = LayoutReplicaParser(use_paddleocr=False)
    monkeypatch.setattr(parser, "_preprocess_tesseract_image", lambda _image: processed_image)

    words = parser._extract_tesseract_words(original_image, 1, 200, 200)

    assert calls == [
        ("original", "--psm 6 -c preserve_interword_spaces=1"),
        ("original", "--psm 11"),
        ("processed", "--psm 11"),
    ]
    assert [word.text for word in words] == [entry[0] for entry in table_entries]
    assert len(words) < len(junk_psm_11)


def test_tesseract_skips_preprocessing_when_original_has_strong_table_coverage(monkeypatch):
    from parsers.layout_replica_parser import LayoutReplicaParser

    entries = []
    for row in range(8):
        top = row * 15
        entries.extend([
            (f"2026-02-{row + 1:02d}", 2, top),
            (f"{row + 1}.00", 100, top),
        ])
    data = {
        "text": [entry[0] for entry in entries],
        "conf": [90] * len(entries),
        "left": [entry[1] for entry in entries],
        "top": [entry[2] for entry in entries],
        "width": [8] * len(entries),
        "height": [8] * len(entries),
    }
    calls = []

    def image_to_data(_image, output_type, config):
        del output_type
        calls.append(config)
        return data

    fake_tesseract = ModuleType("pytesseract")
    fake_tesseract.Output = SimpleNamespace(DICT="dict")
    fake_tesseract.image_to_data = image_to_data
    monkeypatch.setitem(sys.modules, "pytesseract", fake_tesseract)

    parser = LayoutReplicaParser(use_paddleocr=False)
    monkeypatch.setattr(
        parser,
        "_preprocess_tesseract_image",
        lambda _image: pytest.fail("strong table coverage must not add an OCR pass"),
    )
    image = SimpleNamespace(mode="RGB", width=200, height=200)

    words = parser._extract_tesseract_words(image, 1, 200, 200)

    assert len(words) == len(entries)
    assert calls == [
        "--psm 6 -c preserve_interword_spaces=1",
        "--psm 11",
    ]


def test_header_columns_outrank_wider_inferred_columns():
    from parsers.layout_replica_parser import LayoutLine, LayoutPage, LayoutReplicaParser, LayoutWord

    def make_line(page_num, index, words):
        layout_words = [
            LayoutWord(text=text, x0=x0, x1=x1, top=index * 10.0,
                       bottom=index * 10.0 + 8.0, page=page_num, source="pdf-text")
            for text, x0, x1 in words
        ]
        return LayoutLine(page=page_num, index=index, top=index * 10.0,
                          bottom=index * 10.0 + 8.0, center_y=index * 10.0 + 4.0,
                          words=layout_words,
                          text=" ".join(word.text for word in layout_words))

    def make_page(page_num, lines):
        return LayoutPage(page_number=page_num, width=620, height=800, source="ocr",
                          words=[w for line in lines for w in line.words], lines=lines)

    # page 1: no header, transactions only -> positional inference (wide)
    inferred_page = make_page(1, [
        make_line(1, 1, [
            ("29/01/25", 40, 82), ("UPI-SELVI", 100, 160), ("PAY", 170, 195),
            ("REF-991", 210, 255), ("0000024036557328", 270, 360), ("29/01/25", 380, 422),
            ("72.00", 450, 478), ("0.00", 508, 526), ("1,523.00", 570, 596),
        ]),
        make_line(1, 2, [
            ("30/01/25", 40, 82), ("UPI-KUMAR", 100, 160), ("PAY", 170, 195),
            ("REF-992", 210, 255), ("0000024036557329", 270, 360), ("30/01/25", 380, 422),
            ("80.00", 450, 478), ("0.00", 508, 526), ("1,443.00", 570, 596),
        ]),
    ])
    # page 2: proper header with FEWER columns than page 1's inference
    header_page = make_page(2, [
        make_line(2, 1, [
            ("Date", 40, 70), ("Description", 120, 180),
            ("Debit", 330, 360), ("Credit", 410, 445), ("Balance", 500, 545),
        ]),
        make_line(2, 2, [
            ("02/02/25", 40, 82), ("CARD PAYMENT", 120, 200),
            ("25.75", 330, 358), ("1,417.25", 500, 545),
        ]),
    ])

    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [inferred_page, header_page]
    parser._build_table_replica()

    assert [column.header for column in parser.table_columns] == [
        "Date", "Description", "Debit", "Credit", "Balance",
    ]


def test_ocr_header_recovers_date_merges_qualifiers_and_carries_one_schema():
    from parsers.layout_replica_parser import LayoutReplicaParser

    page_one = _make_layout_page(1, [
        _make_layout_line(1, 1, [
            ("Description", 260, 375),
            ("Debits", 660, 710),
            ("(-)", 724, 740),
            ("Credits", 800, 855),
            ("(+)", 870, 886),
            ("Balance", 940, 1025),
        ], source="ocr"),
        _make_layout_line(1, 2, [
            ("2025-11-04", 120, 230),
            ("Deposit", 260, 340),
            ("564.66", 800, 850),
            ("5,564.66", 940, 1010),
        ], source="ocr"),
        _make_layout_line(1, 3, [
            ("2025-11-05", 120, 230),
            ("Payment", 260, 340),
            ("100.00", 660, 710),
            ("5,464.66", 940, 1010),
        ], source="ocr"),
    ], source="ocr")
    page_two = _make_layout_page(2, [
        # This uppercase transaction-like line is deliberately strong enough
        # to look header-ish, but has only one real header-vocabulary token.
        _make_layout_line(2, 1, [
            ("DATE", 120, 180),
            ("PAYMENT", 260, 350),
            ("REFERENCECODE", 400, 560),
            ("ATM", 660, 710),
            ("TOTAL", 940, 1010),
        ], source="ocr"),
        _make_layout_line(2, 2, [
            ("2025-11-06", 120, 230),
            ("Transfer", 260, 340),
            ("200.00", 660, 710),
            ("5,264.66", 940, 1010),
        ], source="ocr"),
    ], source="ocr")

    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [page_one, page_two]
    parser._build_table_replica()

    expected_headers = ("Date", "Description", "Debits (-)", "Credits (+)", "Balance")
    assert tuple(column.header for column in parser.table_columns) == expected_headers
    assert {row.schema_headers for row in parser.table_rows} == {expected_headers}
    assert parser.table_page_summaries[1]["header_line"] == "carried"


def test_ocr_split_header_fragments_merge_without_changing_source_lines():
    from parsers.layout_replica_parser import LayoutReplicaParser

    amount_fragment = _make_layout_line(1, 1, [
        ("Debits", 660, 716),
        ("(-)", 726, 754),
        ("Credits(+)", 802, 905),
        ("Balance", 940, 1023),
    ], source="ocr")
    left_fragment = _make_layout_line(1, 2, [
        ("Date", 117, 172),
        ("Description", 260, 374),
    ], source="ocr")

    # Mimic a perspective-skewed header whose left and right fragments overlap
    # vertically but were emitted as two OCR lines.
    for line, top, bottom in (
        (amount_fragment, 30.0, 62.0),
        (left_fragment, 52.0, 78.0),
    ):
        line.top = top
        line.bottom = bottom
        line.center_y = (top + bottom) / 2.0
        for word in line.words:
            word.top = top
            word.bottom = bottom

    page = _make_layout_page(1, [amount_fragment, left_fragment], source="ocr")
    page.width = 1200
    parser = LayoutReplicaParser(use_ocr=False)

    header = parser._detect_table_header_line(page)
    assert header is not None
    columns = parser._columns_from_header_line(header, page.width)

    assert [column.header for column in columns] == [
        "Date", "Description", "Debits (-)", "Credits (+)", "Balance",
    ]
    assert [line.text for line in page.lines] == [
        "Debits (-) Credits(+) Balance",
        "Date Description",
    ]


@pytest.mark.parametrize(
    "credit_rows",
    ({1, 3}, {3}, set()),
    ids=("two-sided", "single-credit", "debit-only"),
)
def test_ocr_date_anchors_recover_shifted_rows_and_amount_columns(credit_rows):
    from parsers.layout_replica_parser import LayoutReplicaParser

    lines = [
        _make_layout_line(1, 1, [
            ("Date", 120, 180),
            ("Description", 260, 375),
            ("Debits", 660, 716),
            ("(-)", 726, 754),
            ("Credits", 802, 868),
            ("(+)", 877, 905),
            ("Balance", 940, 1023),
        ], source="ocr"),
    ]
    expected = []
    for row_number in range(4):
        date = f"2026-01-{row_number + 1:02d}"
        description = f"Transaction {row_number + 1}"
        credit = f"{200 + row_number}.00" if row_number in credit_rows else ""
        debit = "" if credit else f"{100 + row_number}.00"
        balance = f"{1000 - row_number}.00"
        date_line_index = 2 + (row_number * 2)
        amount_line_index = date_line_index + 1
        lines.append(_make_layout_line(1, date_line_index, [
            # The shifted box overlaps Description more than the carried Date
            # bounds, matching the failure seen on perspective-skewed scans.
            (date, 180, 290),
            (description, 305, 450),
            # A final description token can overlap the old Debit bounds after
            # page drift; lexical text before the observed amount band remains
            # part of Description.
            ("Tail", 630, 700),
        ], source="ocr"))
        amount_words = []
        if debit:
            amount_words.append((debit, 760, 840))
        if credit:
            # Its center is to the right of the old Balance x0, but still left
            # of the observed balance band. The page-local amount clusters must
            # keep it in Credit.
            amount_words.append((credit, 910, 980))
        amount_words.append((balance, 1000, 1090))
        lines.append(_make_layout_line(
            1,
            amount_line_index,
            amount_words,
            source="ocr",
        ))
        expected.append([date, f"{description} Tail", debit, credit, balance])

    page = _make_layout_page(1, lines, source="ocr")
    page.width = 1200
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [page]
    parser._build_table_replica()

    assert [column.header for column in parser.table_columns] == [
        "Date", "Description", "Debits (-)", "Credits (+)", "Balance",
    ]
    assert [row.values for row in parser.table_rows] == expected


def test_ocr_split_three_column_header_uses_date_amount_anchors():
    from parsers.layout_replica_parser import LayoutReplicaParser

    amount_header = _make_layout_line(1, 1, [
        ("Amount", 450, 484),
    ], source="ocr")
    left_header = _make_layout_line(1, 2, [
        ("Date", 60, 82),
        ("Description", 140, 192),
    ], source="ocr")
    lines = [amount_header, left_header]

    def place_line(line, top, bottom):
        line.top = top
        line.bottom = bottom
        line.center_y = (top + bottom) / 2.0
        for word in line.words:
            word.top = top
            word.bottom = bottom

    # The one-word right header and two-word left header overlap visually but
    # are split by OCR, matching a lightly rotated card-statement scan.
    place_line(amount_header, 90.0, 103.0)
    place_line(left_header, 99.0, 116.0)

    expected = []
    for row_number in range(4):
        date = f"06/{row_number + 2:02d}"
        description = f"CARD ITEM {row_number + 1}"
        amount = f"-{20 + row_number}.00"
        amount_line = _make_layout_line(
            1,
            3 + (row_number * 2),
            [(amount, 450, 485)],
            source="ocr",
        )
        date_line = _make_layout_line(
            1,
            4 + (row_number * 2),
            [(date, 60, 84), (description, 140, 260)],
            source="ocr",
        )
        top = 125.0 + (row_number * 40.0)
        place_line(amount_line, top, top + 12.0)
        place_line(date_line, top + 10.0, top + 28.0)
        lines.extend((amount_line, date_line))
        expected.append([date, description, amount])

    page = _make_layout_page(1, lines, source="ocr")
    parser = LayoutReplicaParser(use_ocr=False)
    parser.pages = [page]
    parser._build_table_replica()

    assert [column.header for column in parser.table_columns] == [
        "Date", "Description", "Amount",
    ]
    assert [row.values for row in parser.table_rows] == expected


def test_date_detection_covers_western_formats():
    from parsers.layout_replica_parser import LayoutReplicaParser

    accepted = ["06/11", "06/03/2026", "3 Jun 2026", "30 June 2026", "Jun 3, 2026", "03.06.2026", "03-06-2026"]
    rejected = ["1.5", "1,234.56", "2,100.00", "CHECK 1042", "", "2026"]
    for value in accepted:
        assert LayoutReplicaParser._looks_like_date_value(value), value
    for value in rejected:
        assert not LayoutReplicaParser._looks_like_date_value(value), value


def test_amount_detection_covers_western_formats():
    from parsers.layout_replica_parser import LayoutReplicaParser

    accepted = ["1,234.56", "£2,480.00", "€112,40", "$300.00", "-15.99", "(45.00)", "1.234,56", "2480", "0.84", "-5,711.97."]
    rejected = ["3 Jun 2026", "06/11", "CHECK", "", "-", "£"]
    for value in accepted:
        assert LayoutReplicaParser._looks_like_amount_value(value), value
    for value in rejected:
        assert not LayoutReplicaParser._looks_like_amount_value(value), value


def test_amount_headers_cover_western_conventions():
    from parsers.layout_replica_parser import LayoutReplicaParser

    accepted = ["Debit", "Withdrawals", "Money out", "Money in", "Af", "Bij", "Saldo", "Soll", "Haben", "Amount"]
    rejected = ["Description", "Omschrijving", "Date", "Reference"]
    for header in accepted:
        assert LayoutReplicaParser._is_amount_header(header), header
    for header in rejected:
        assert not LayoutReplicaParser._is_amount_header(header), header


def test_word_snaps_to_column_holding_most_of_its_width():
    from parsers.layout_replica_parser import LayoutReplicaParser, LayoutWord, TableColumn

    columns = [
        TableColumn(header="DATE", x0=10.0, x1=90.0, left=0.0, right=100.0),
        TableColumn(header="VALUE DT", x0=100.0, x1=108.0, left=100.0, right=110.0),
        TableColumn(header="DESCRIPTION", x0=115.0, x1=290.0, left=110.0, right=300.0),
    ]
    # Proportionally-estimated OCR token box straddles the narrow VALUE DT
    # column: its center (x=100.0) lands there, but 70pt of its width lies in
    # DATE vs 10pt in VALUE DT. Most-of-width must win over center.
    word = LayoutWord(text="03/09/18", x0=30.0, x1=170.0, top=0.0, bottom=10.0, page=1, source="ocr")

    assert LayoutReplicaParser._table_column_index_for_word(word, columns) == 0


def test_word_fully_inside_a_column_is_unaffected_by_snapping():
    from parsers.layout_replica_parser import LayoutReplicaParser, LayoutWord, TableColumn

    columns = [
        TableColumn(header="DATE", x0=10.0, x1=90.0, left=0.0, right=100.0),
        TableColumn(header="VALUE DT", x0=100.0, x1=108.0, left=100.0, right=110.0),
        TableColumn(header="DESCRIPTION", x0=115.0, x1=290.0, left=110.0, right=300.0),
    ]
    word = LayoutWord(text="UPI-SELVI", x0=120.0, x1=160.0, top=0.0, bottom=10.0, page=1, source="ocr")

    assert LayoutReplicaParser._table_column_index_for_word(word, columns) == 2


def test_word_just_past_last_column_edge_still_lands_in_last_column():
    from parsers.layout_replica_parser import LayoutReplicaParser, LayoutWord, TableColumn

    columns = [
        TableColumn(header="DATE", x0=10.0, x1=90.0, left=0.0, right=100.0),
        TableColumn(header="BALANCE", x0=110.0, x1=290.0, left=100.0, right=300.0),
    ]
    word = LayoutWord(text="1,523.00", x0=302.0, x1=330.0, top=0.0, bottom=10.0, page=1, source="ocr")

    assert LayoutReplicaParser._table_column_index_for_word(word, columns) == 1


def _numeric_refinement_fixture():
    from parsers.layout_replica_parser import LayoutPage, LayoutReplicaParser, LayoutWord, TableColumn

    columns = [
        TableColumn("Date", 40.0, 100.0, 30.0, 110.0),
        TableColumn("Description", 120.0, 300.0, 110.0, 310.0),
        TableColumn("Debits (-)", 330.0, 390.0, 310.0, 400.0),
        TableColumn("Credits (+)", 420.0, 480.0, 400.0, 490.0),
        TableColumn("Balance", 510.0, 580.0, 490.0, 600.0),
    ]
    source_balance = LayoutWord(
        "3,225.65",
        510.0,
        575.0,
        80.0,
        92.0,
        1,
        "ocr",
        0.93,
    )
    word_buckets = [
        [[], [], [], [], [LayoutWord("-3,000.00", 510, 580, 40, 52, 1, "ocr", 0.95)]],
        [[], [], [LayoutWord("225.65", 330, 385, 80, 92, 1, "ocr", 0.96)], [], [source_balance]],
    ]
    values = [
        ["2026-01-01", "Opening", "", "", "-3,000.00"],
        ["2026-01-02", "Payment", "225.65", "", "3,225.65"],
    ]
    page = LayoutPage(
        page_number=1,
        width=620.0,
        height=800.0,
        words=[word for row in word_buckets for cell in row for word in cell],
        lines=[],
        source="ocr",
    )
    parser = LayoutReplicaParser(use_ocr=True, use_paddleocr=True)
    parser._pdf_path = "statement.pdf"
    return parser, page, columns, word_buckets, values, source_balance


def test_cell_local_numeric_ocr_recovers_sign_without_changing_exact_copy(monkeypatch):
    parser, page, columns, word_buckets, values, source_balance = _numeric_refinement_fixture()

    def direct_read(*args):
        dpi = args[-2]
        half_height = args[-1]
        assert dpi in {150, 250}
        assert half_height == 30.0
        return "-3,225.65"

    monkeypatch.setattr(parser, "_read_numeric_cell_crop", direct_read)
    parser._refine_ocr_numeric_cells(page, columns, word_buckets, values)

    assert values[1][4] == "-3,225.65"
    assert source_balance.text == "3,225.65"
    assert page.words[-1] is source_balance
    assert parser._numeric_ocr_refinement_count == 1
    assert parser._numeric_ocr_unresolved_count == 0


@pytest.mark.parametrize("fresh_run", range(2))
def test_cell_local_numeric_ocr_requires_repeatable_negative_votes_across_render_dpis(
    monkeypatch,
    fresh_run,
):
    del fresh_run
    parser, page, columns, word_buckets, values, source_balance = _numeric_refinement_fixture()
    calls = []

    def pymupdf_like_read(*args):
        dpi = args[-2]
        half_height = args[-1]
        calls.append((half_height, dpi))
        if half_height == 30.0:
            return {
                150: "3,225.65",
                165: "3,225.65",
                200: "3,225.65",
                220: "-3,225.65",
            }.get(dpi)
        if half_height == 35.0:
            return {
                150: "3,225.65",
                180: "-3,225.65",
                165: "-3,225.65",
            }.get(dpi)
        return None

    monkeypatch.setattr(parser, "_read_numeric_cell_crop", pymupdf_like_read)
    parser._refine_ocr_numeric_cells(page, columns, word_buckets, values)

    assert values[1][4] == "-3,225.65"
    assert source_balance.text == "3,225.65"
    assert (35.0, 180) in calls
    assert (35.0, 165) in calls
    assert parser._numeric_ocr_refinement_count == 1
    assert parser._numeric_ocr_unresolved_count == 0


def test_cell_local_numeric_ocr_disagreement_does_not_infer_sign(monkeypatch):
    parser, page, columns, word_buckets, values, source_balance = _numeric_refinement_fixture()

    def disagreeing_read(*args):
        dpi = args[-2]
        return "-3,225.65" if dpi == 150 else "3,225.65"

    monkeypatch.setattr(parser, "_read_numeric_cell_crop", disagreeing_read)
    parser._refine_ocr_numeric_cells(page, columns, word_buckets, values)
    parser.pages = [page]
    parser._build_metadata()
    parser.quality_report = parser._build_quality_report()

    assert values[1][4] == "3,225.65"
    assert source_balance.text == "3,225.65"
    assert parser._numeric_ocr_refinement_count == 0
    assert parser._numeric_ocr_unresolved_count == 1
    assert parser.get_quality_report()["review_required"] is True
    assert parser.get_quality_report()["numeric_ocr_unresolved_count"] == 1


def test_cell_local_numeric_ocr_never_accepts_different_digits(monkeypatch):
    parser, page, columns, word_buckets, values, source_balance = _numeric_refinement_fixture()
    monkeypatch.setattr(
        parser,
        "_read_numeric_cell_crop",
        lambda *_args: "-3,226.65",
    )

    parser._refine_ocr_numeric_cells(page, columns, word_buckets, values)

    assert values[1][4] == "3,225.65"
    assert source_balance.text == "3,225.65"
    assert parser._numeric_ocr_refinement_count == 0
    assert parser._numeric_ocr_unresolved_count == 1


def test_strict_money_syntax_requires_unambiguous_separators():
    from parsers.layout_replica_parser import LayoutReplicaParser

    accepted = ("1,498.03", "1.498,03", "-10,889.97", "₹1,23,456.78", "2480")
    rejected = ("1.498.03", "3.225.65", "10,889,97", "-5,711.97.")

    assert all(LayoutReplicaParser._strict_money_value(value) is not None for value in accepted)
    assert all(LayoutReplicaParser._strict_money_value(value) is None for value in rejected)


def test_numeric_crop_keeps_source_negative_when_repairing_only_punctuation():
    from parsers.layout_replica_parser import LayoutReplicaParser

    results = [{
        "text": "2,313.62",
        "confidence": 0.97,
        "bbox": (40.0, 30.0, 140.0, 70.0),
    }]

    assert LayoutReplicaParser._numeric_candidate_from_crop_results(
        results,
        "-2.313.62",
        crop_width=180,
        crop_height=100,
    ) == "-2,313.62"
