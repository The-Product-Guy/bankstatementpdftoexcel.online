"""Slow production-path fidelity gates for image-only bank statements.

Descriptions may contain harmless OCR character or spacing differences, but a
visual transaction must not disappear, move out of order, change schema, shift
columns, or alter any printed debit, credit, balance, sign, or separator.
"""

from __future__ import annotations

import csv
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re

from openpyxl import load_workbook
import pytest


SYNTHETIC_DIR = Path(__file__).parent / "data" / "synthetic"
EXPECTED_HEADERS = ("date", "description", "debits", "credits", "balance")
FIXTURES = (
    ("stmt_canada_001", 46),
    ("stmt_canada_002", 27),
    ("stmt_canada_003", 17),
    ("stmt_canada_004", 45),
    ("stmt_canada_005", 13),
)


def _read_truth(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _normalize_header(value: object) -> str:
    return re.sub(r"[^a-z]+", "", str(value or "").casefold())


def _normalize_date(value: object) -> str:
    """Normalize separators and a common OCR O/0 substitution in ISO dates."""
    text = str(value or "").strip().upper().replace("O", "0")
    match = re.search(r"(20[0-9]{2})\D([0-9]{2})\D([0-9]{2})", text)
    return "-".join(match.groups()) if match else text


def _nonzero(value: str) -> bool:
    return Decimal(value or "0") != 0


def _money_tokens(value: object) -> list[str]:
    """Count amount-shaped groups while tolerating OCR punctuation inside one group."""
    return re.findall(
        r"[-(]?\d[\d,.]*[.,]\d{2}(?:CR|DR)?\)?",
        str(value or ""),
        re.I,
    )


def _money_to_signed_cents(value: object) -> int:
    """Parse a displayed Western money value without forgiving digit/sign OCR errors.

    The final comma or full stop followed by exactly two digits is the decimal
    separator. Earlier commas/full stops/spaces are grouping separators, so OCR
    variants such as ``1,234.56``, ``1.234,56``, and ``1.234.56`` canonicalize
    identically. Only separator glyphs are ignored: every digit and an explicit
    negative sign still affects the result.
    """
    text = str(value or "").strip().upper()
    text = text.translate(str.maketrans({"−": "-", "–": "-", "—": "-", "﹣": "-"}))

    parenthesized = text.startswith("(") and text.endswith(")")
    if parenthesized:
        text = text[1:-1].strip()

    suffix_sign = None
    suffix_match = re.search(r"\s*(CR|DR)\s*$", text)
    if suffix_match:
        suffix_sign = -1 if suffix_match.group(1) == "DR" else 1
        text = text[: suffix_match.start()].strip()

    explicit_minus = text.startswith("-") or text.endswith("-")
    explicit_plus = text.startswith("+")
    if text.startswith(("+", "-")):
        text = text[1:].strip()
    if text.endswith("-"):
        text = text[:-1].strip()

    # Currency marks and grouping spaces do not change the represented cents.
    text = re.sub(r"[$£€¥₹]", "", text).strip()
    if not re.fullmatch(r"\d[\d.,'’\s]*", text):
        raise ValueError(f"not an unambiguous money value: {value!r}")

    decimal_match = re.search(r"[.,](\d{2})(?:[.,\s]*)$", text)
    if not decimal_match:
        raise ValueError(f"money value does not preserve two printed cents: {value!r}")

    integer_digits = re.sub(r"\D", "", text[: decimal_match.start()])
    if not integer_digits:
        raise ValueError(f"money value lost its integer digits: {value!r}")

    cents = int(integer_digits) * 100 + int(decimal_match.group(1))
    if suffix_sign == 1 and (parenthesized or explicit_minus):
        raise ValueError(f"money value has conflicting signs: {value!r}")
    if suffix_sign == -1 and explicit_plus:
        raise ValueError(f"money value has conflicting signs: {value!r}")
    return -cents if parenthesized or explicit_minus or suffix_sign == -1 else cents


def _truth_to_cents(value: str) -> int:
    """Round generator float artifacts to the two decimals printed in the PDF."""
    return int(
        (Decimal(value or "0").quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100)
        .to_integral_exact()
    )


def _truth_to_printed_money(value: str) -> str:
    """Match the two-decimal, comma-grouped format printed by the generator."""
    amount = Decimal(value or "0").quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{amount:,.2f}"


@pytest.mark.parametrize(
    ("printed", "expected_cents"),
    (
        ("1,234.56", 123456),
        ("1.234,56", 123456),
        ("1.234.56", 123456),
        ("10,889,97", 1088997),
        ("1,234.56.", 123456),
        ("£1 234,56", 123456),
        ("(1,234.56)", -123456),
        ("1,234.56 DR", -123456),
        ("1,234.56 CR", 123456),
    ),
)
def test_money_canonicalization_only_ignores_separator_variants(
    printed: str,
    expected_cents: int,
) -> None:
    assert _money_to_signed_cents(printed) == expected_cents


@pytest.mark.parametrize(
    "printed",
    ("123456", "1,23O.56", "1.234", "-1.00 CR", "+1.00 DR"),
)
def test_money_canonicalization_rejects_missing_digits_or_conflicting_signs(
    printed: str,
) -> None:
    with pytest.raises(ValueError):
        _money_to_signed_cents(printed)


@pytest.mark.ocr
@pytest.mark.parametrize("quality", ("standard", "high"))
@pytest.mark.parametrize(("fixture_name", "expected_row_count"), FIXTURES)
def test_scanned_statement_preserves_rows_columns_and_workbook(
    fixture_name: str,
    expected_row_count: int,
    quality: str,
    tmp_path: Path,
) -> None:
    from parsers.layout_replica_parser import create_layout_replica_parser

    pdf_path = SYNTHETIC_DIR / f"{fixture_name}_scanned.pdf"
    truth_path = SYNTHETIC_DIR / f"{fixture_name}_truth.csv"
    truth_rows = _read_truth(truth_path)

    assert len(truth_rows) == expected_row_count

    # This is the same layout-replica factory used by conversion jobs, with
    # the primary ONNX OCR backend enabled and Tesseract left as its fallback.
    parser = create_layout_replica_parser(
        quality=quality,
        use_ocr=True,
        use_paddleocr=True,
    )
    parser.parse(str(pdf_path), pdf_path.name)

    parsed_rows = parser.table_rows
    assert len(parsed_rows) == expected_row_count, (
        f"{fixture_name}: expected {expected_row_count} visual transaction rows, "
        f"extracted {len(parsed_rows)}"
    )

    expected_dates = [row["Date"] for row in truth_rows]
    parsed_dates = [_normalize_date(row.values[0]) for row in parsed_rows]
    assert parsed_dates == expected_dates, (
        f"{fixture_name}: transaction rows are missing, duplicated, or out of "
        "source order"
    )

    schemas = {
        tuple(_normalize_header(header) for header in row.schema_headers)
        for row in parsed_rows
    }
    assert schemas == {EXPECTED_HEADERS}, (
        f"{fixture_name}: expected one stable five-column schema, got {schemas}"
    )
    assert tuple(_normalize_header(column.header) for column in parser.table_columns) == EXPECTED_HEADERS
    assert all(len(row.values) == len(EXPECTED_HEADERS) for row in parsed_rows)
    assert all(str(row.values[1] or "").strip() for row in parsed_rows), (
        f"{fixture_name}: a transaction row lost its description cell"
    )
    assert all(str(row.values[4] or "").strip() for row in parsed_rows), (
        f"{fixture_name}: a transaction row lost its running-balance cell"
    )
    assert all(len(_money_tokens(row.values[4])) == 1 for row in parsed_rows), (
        f"{fixture_name}: a transaction amount was merged into or duplicated in "
        "the running-balance cell"
    )

    transaction_matches = 0
    balance_matches = 0
    numeric_failures: list[str] = []
    display_failures: list[str] = []
    for row_number, (truth, parsed) in enumerate(zip(truth_rows, parsed_rows), 1):
        expected_debit = _nonzero(truth["Withdrawal_Amount"])
        expected_credit = _nonzero(truth["Deposit_Amount"])
        assert expected_debit != expected_credit, (
            f"{fixture_name} truth row {row_number} must contain exactly one "
            "withdrawal or deposit"
        )

        # OCR spelling is deliberately outside this structural gate, but every
        # printed transaction amount must land in exactly one transaction-side
        # cell. Skipping empty/misplaced rows would hide column corruption.
        parsed_debit = bool(str(parsed.values[2] or "").strip())
        parsed_credit = bool(str(parsed.values[3] or "").strip())
        assert (parsed_debit, parsed_credit) == (expected_debit, expected_credit), (
            f"{fixture_name} row {row_number} ({truth['Date']}): detected amount "
            f"is in the wrong transaction column: debit={parsed.values[2]!r}, "
            f"credit={parsed.values[3]!r}"
        )
        assert (
            len(_money_tokens(parsed.values[2])),
            len(_money_tokens(parsed.values[3])),
        ) == (int(expected_debit), int(expected_credit)), (
            f"{fixture_name} row {row_number} ({truth['Date']}): transaction "
            "amount was lost, duplicated, or is not recognizable in its cell"
        )

        # Compare the exact signed cents, not merely whether an amount-shaped
        # token landed in the expected column. Debit/credit placement supplies
        # the transaction sign; an OCR-invented minus sign reverses it and fails.
        truth_amount = (
            truth["Withdrawal_Amount"] if expected_debit else truth["Deposit_Amount"]
        )
        expected_transaction_cents = _truth_to_cents(truth_amount)
        if expected_debit:
            expected_transaction_cents = -expected_transaction_cents
        parsed_amount = parsed.values[2] if expected_debit else parsed.values[3]
        expected_transaction_display = _truth_to_printed_money(truth_amount)
        if str(parsed_amount or "").strip() != expected_transaction_display:
            display_failures.append(
                f"row {row_number} {truth['Date']} transaction display: "
                f"expected {expected_transaction_display!r}, got {parsed_amount!r}"
            )
        try:
            parsed_amount_cents = _money_to_signed_cents(parsed_amount)
            parsed_transaction_cents = (
                -parsed_amount_cents if expected_debit else parsed_amount_cents
            )
        except ValueError as exc:
            parsed_transaction_cents = None
            numeric_failures.append(
                f"row {row_number} {truth['Date']} transaction: "
                f"expected {expected_transaction_cents} cents, got "
                f"{parsed_amount!r} ({exc})"
            )
        if parsed_transaction_cents == expected_transaction_cents:
            transaction_matches += 1
        elif parsed_transaction_cents is not None:
            numeric_failures.append(
                f"row {row_number} {truth['Date']} transaction: "
                f"expected {expected_transaction_cents} cents, got "
                f"{parsed_transaction_cents} from {parsed_amount!r}"
            )

        expected_balance_cents = _truth_to_cents(truth["Closing_Balance"])
        expected_balance_display = _truth_to_printed_money(truth["Closing_Balance"])
        if str(parsed.values[4] or "").strip() != expected_balance_display:
            display_failures.append(
                f"row {row_number} {truth['Date']} balance display: expected "
                f"{expected_balance_display!r}, got {parsed.values[4]!r}"
            )
        try:
            parsed_balance_cents = _money_to_signed_cents(parsed.values[4])
        except ValueError as exc:
            parsed_balance_cents = None
            numeric_failures.append(
                f"row {row_number} {truth['Date']} balance: expected "
                f"{expected_balance_cents} cents, got {parsed.values[4]!r} ({exc})"
            )
        if parsed_balance_cents == expected_balance_cents:
            balance_matches += 1
        elif parsed_balance_cents is not None:
            numeric_failures.append(
                f"row {row_number} {truth['Date']} balance: expected "
                f"{expected_balance_cents} cents, got {parsed_balance_cents} "
                f"from {parsed.values[4]!r}"
            )

    assert not numeric_failures, (
        f"{fixture_name} ({quality}) exact signed-cent fidelity failed: "
        f"transaction amounts {transaction_matches}/{expected_row_count} "
        f"({transaction_matches / expected_row_count:.2%}); running balances "
        f"{balance_matches}/{expected_row_count} "
        f"({balance_matches / expected_row_count:.2%}). Examples: "
        + "; ".join(numeric_failures[:8])
    )
    assert not display_failures, (
        f"{fixture_name} ({quality}) printed money fidelity failed. Examples: "
        + "; ".join(display_failures[:8])
    )

    output_path = tmp_path / f"{fixture_name}.xlsx"
    parser.write_excel(str(output_path))
    workbook = load_workbook(output_path, data_only=False)
    table_sheets = [name for name in workbook.sheetnames if name.startswith("Table_Data")]
    assert table_sheets == ["Table_Data"], (
        f"{fixture_name}: one source table must serialize to one worksheet"
    )

    sheet = workbook["Table_Data"]
    workbook_headers = tuple(
        _normalize_header(sheet.cell(1, column).value)
        for column in range(1, len(EXPECTED_HEADERS) + 1)
    )
    assert workbook_headers == EXPECTED_HEADERS
    assert sheet.max_row == expected_row_count + 1
    assert sheet.max_column == len(EXPECTED_HEADERS)

    workbook_rows = [
        [str(sheet.cell(row_number, column).value or "") for column in range(1, 6)]
        for row_number in range(2, sheet.max_row + 1)
    ]
    parser_rows = [[str(value or "") for value in row.values] for row in parsed_rows]
    assert workbook_rows == parser_rows, (
        f"{fixture_name}: XLSX serialization changed, dropped, or shifted table cells"
    )
    workbook.close()


@pytest.mark.ocr
@pytest.mark.parametrize("quality", ("standard", "high"))
def test_rotated_scanned_card_preserves_three_column_rows(
    quality: str,
    tmp_path: Path,
) -> None:
    """A split 1+2 word OCR header must still form Date/Description/Amount."""
    from pdf2image import convert_from_path
    from PIL import Image
    from parsers.layout_replica_parser import create_layout_replica_parser

    source_pdf = SYNTHETIC_DIR / "us-card.pdf"
    scanned_pdf = tmp_path / "us-card-rotated-scan.pdf"
    rendered_pages = convert_from_path(
        str(source_pdf),
        dpi=150,
        first_page=1,
        last_page=1,
        fmt="png",
    )
    assert len(rendered_pages) == 1
    rotated = rendered_pages[0].convert("RGB").rotate(
        2.0,
        resample=Image.Resampling.BICUBIC,
        expand=False,
        fillcolor="white",
    )
    rotated.save(scanned_pdf, "PDF", resolution=150)

    parser = create_layout_replica_parser(
        quality=quality,
        use_ocr=True,
        use_paddleocr=True,
    )
    parser.parse(str(scanned_pdf), scanned_pdf.name)

    expected_headers = ("date", "description", "amount")
    expected_dates = [
        "06/02", "06/04", "06/07", "06/11",
        "06/15", "06/19", "06/23", "06/28",
    ]
    expected_amounts = [
        "-15.99", "-64.20", "-412.50", "2,100.00",
        "-133.07", "-27.80", "-89.99", "-12.99",
    ]
    assert tuple(_normalize_header(column.header) for column in parser.table_columns) == expected_headers
    assert {
        tuple(_normalize_header(header) for header in row.schema_headers)
        for row in parser.table_rows
    } == {expected_headers}
    assert [_normalize_date(row.values[0]) for row in parser.table_rows] == expected_dates
    assert all(len(row.values) == 3 for row in parser.table_rows)
    assert all(str(row.values[1] or "").strip() for row in parser.table_rows)
    assert [str(row.values[2] or "").strip() for row in parser.table_rows] == expected_amounts

    output_path = tmp_path / "us-card-rotated-scan.xlsx"
    parser.write_excel(str(output_path))
    workbook = load_workbook(output_path, data_only=False)
    assert workbook.sheetnames == ["Exact_Copy", "Table_Data", "Full_Text"]
    assert workbook["Table_Data"].max_row == len(expected_dates) + 1
    assert workbook["Table_Data"].max_column == len(expected_headers)
    workbook.close()
