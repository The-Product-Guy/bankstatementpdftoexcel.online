#!/usr/bin/env python3
"""Fidelity gates for the committed US/UK/EU statement fixtures.

Fixtures are synthetic (tools/make_synthetic_statements.py). Each family
varies column sets, date formats, and amount conventions. These tests back
the /convert/<bank> marketing claims: no page ships for a format family the
production layout parser cannot reproduce cell-for-cell.
"""
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

FIXTURES = Path(__file__).parent / "data" / "synthetic"
REPO_ROOT = Path(__file__).parent.parent


# Expected values are the source data in tools/make_synthetic_statements.py.
CASES = [
    pytest.param(
        "us-checking.pdf",
        ["FIRST DEMO BANK N.A.", "Checking Statement 06/01/2026 - 06/30/2026"],
        ["Date", "Description", "Withdrawals", "Deposits", "Balance"],
        [
            ["06/01/2026", "BEGINNING BALANCE", "", "", "8,410.22"],
            ["06/03/2026", "DIRECT DEPOSIT - EMPLOYER PAYROLL", "", "3,250.00", "11,660.22"],
            ["06/05/2026", "CARD PURCHASE - GROCERY STORE", "182.45", "", "11,477.77"],
            ["06/09/2026", "ACH PAYMENT - UTILITIES", "240.10", "", "11,237.67"],
            ["06/12/2026", "ATM WITHDRAWAL", "300.00", "", "10,937.67"],
            ["06/18/2026", "CHECK 1042", "1,150.00", "", "9,787.67"],
            ["06/24/2026", "CARD PURCHASE - GAS STATION", "58.90", "", "9,728.77"],
            ["06/30/2026", "INTEREST PAYMENT", "", "1.12", "9,729.89"],
        ],
        id="us-checking",
    ),
    pytest.param(
        "us-card.pdf",
        ["DEMO BANK CARD SERVICES", "Cardmember Statement - June 2026"],
        ["Date", "Description", "Amount"],
        [
            ["06/02", "ONLINE SUBSCRIPTION SERVICE", "-15.99"],
            ["06/04", "RESTAURANT - DOWNTOWN", "-64.20"],
            ["06/07", "AIRLINE TICKET", "-412.50"],
            ["06/11", "PAYMENT RECEIVED - THANK YOU", "2,100.00"],
            ["06/15", "GROCERY STORE", "-133.07"],
            ["06/19", "RIDESHARE", "-27.80"],
            ["06/23", "HARDWARE STORE", "-89.99"],
            ["06/28", "STREAMING SERVICE", "-12.99"],
        ],
        id="us-card",
    ),
    pytest.param(
        "uk-current.pdf",
        ["DEMO BANK UK PLC", "Current Account Statement 1 Jun - 30 Jun 2026"],
        ["Date", "Description", "Money out", "Money in", "Balance"],
        [
            ["1 Jun 2026", "BALANCE BROUGHT FORWARD", "", "", "£4,120.55"],
            ["3 Jun 2026", "FASTER PAYMENT - SALARY", "", "£2,480.00", "£6,600.55"],
            ["5 Jun 2026", "DIRECT DEBIT - COUNCIL TAX", "£165.00", "", "£6,435.55"],
            ["9 Jun 2026", "CARD PAYMENT - SUPERMARKET", "£82.14", "", "£6,353.41"],
            ["12 Jun 2026", "STANDING ORDER - RENT", "£1,234.56", "", "£5,118.85"],
            ["18 Jun 2026", "CARD PAYMENT - RAIL TICKETS", "£45.30", "", "£5,073.55"],
            ["24 Jun 2026", "DIRECT DEBIT - ENERGY", "£98.77", "", "£4,974.78"],
            ["30 Jun 2026", "INTEREST", "", "£0.84", "£4,975.62"],
        ],
        id="uk-current",
    ),
    pytest.param(
        "eu-giro.pdf",
        ["DEMO BANK EUROPE", "Rekeningafschrift 01-06-2026 t/m 30-06-2026"],
        ["Datum", "Omschrijving", "Af", "Bij", "Saldo"],
        [
            ["01-06-2026", "BEGINSALDO", "", "", "6.410,22"],
            ["03-06-2026", "SALARIS WERKGEVER", "", "2.850,00", "9.260,22"],
            ["05-06-2026", "INCASSO ENERGIE", "112,40", "", "9.147,82"],
            ["09-06-2026", "PINBETALING SUPERMARKT", "94,18", "", "9.053,64"],
            ["12-06-2026", "HUUR JUNI", "1.234,56", "", "7.819,08"],
            ["18-06-2026", "OVERBOEKING SPAARREKENING", "500,00", "", "7.319,08"],
            ["24-06-2026", "PINBETALING TANKSTATION", "68,45", "", "7.250,63"],
            ["30-06-2026", "RENTE", "", "1,05", "7.251,68"],
        ],
        id="eu-giro",
    ),
]


def _normalized_line(value):
    return " ".join(str(value or "").split())


def _sheet_cells(sheet):
    return [
        [(cell.value, cell.data_type) for cell in row]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=sheet.max_row,
            min_col=1,
            max_col=sheet.max_column,
        )
    ]


@pytest.mark.parametrize("fixture,bank_lines,expected_headers,expected_rows", CASES)
def test_western_format_replicated(
    tmp_path,
    fixture,
    bank_lines,
    expected_headers,
    expected_rows,
):
    from parsers.layout_replica_parser import create_layout_replica_parser

    parser = create_layout_replica_parser(use_ocr=False)
    parser.parse(str(FIXTURES / fixture), fixture)

    assert [column.header for column in parser.table_columns] == expected_headers
    assert [row.values for row in parser.table_rows] == expected_rows

    expected_source_lines = [
        *bank_lines,
        " ".join(expected_headers),
        *(" ".join(cell for cell in row if cell) for row in expected_rows),
    ]
    assert len(expected_source_lines) == 11
    assert [_normalized_line(line.text) for page in parser.pages for line in page.lines] == expected_source_lines

    output_path = tmp_path / f"{Path(fixture).stem}.xlsx"
    parser.write_excel(str(output_path))
    workbook = load_workbook(output_path)

    assert workbook.sheetnames == ["Exact_Copy", "Table_Data", "Full_Text"]

    table_sheet = workbook["Table_Data"]
    serialized_table = [
        [
            "" if table_sheet.cell(row=row_idx, column=col_idx).value is None
            else str(table_sheet.cell(row=row_idx, column=col_idx).value)
            for col_idx in range(1, len(expected_headers) + 1)
        ]
        for row_idx in range(1, len(expected_rows) + 2)
    ]
    assert table_sheet.max_row == len(expected_rows) + 1
    assert table_sheet.max_column == len(expected_headers)
    assert serialized_table == [expected_headers, *expected_rows]

    exact_copy = workbook["Exact_Copy"]
    assert exact_copy.cell(1, 1).value == "Page 1"
    exact_copy_lines = [
        _normalized_line(" ".join(str(cell.value) for cell in exact_copy[row_idx] if cell.value is not None))
        for row_idx in range(2, exact_copy.max_row + 1)
    ]
    assert exact_copy.max_row == len(expected_source_lines) + 1
    assert exact_copy_lines == expected_source_lines

    full_text = workbook["Full_Text"]
    assert [full_text.cell(1, col_idx).value for col_idx in range(1, 5)] == [
        "Page",
        "Line",
        "Source",
        "Text",
    ]
    assert full_text.max_row == len(expected_source_lines) + 1
    assert [full_text.cell(row_idx, 1).value for row_idx in range(2, full_text.max_row + 1)] == [
        "1"
    ] * len(expected_source_lines)
    assert [full_text.cell(row_idx, 2).value for row_idx in range(2, full_text.max_row + 1)] == [
        str(line_idx) for line_idx in range(1, len(expected_source_lines) + 1)
    ]
    assert {full_text.cell(row_idx, 3).value for row_idx in range(2, full_text.max_row + 1)} == {
        "pdf-text"
    }
    assert [
        _normalized_line(full_text.cell(row_idx, 4).value)
        for row_idx in range(2, full_text.max_row + 1)
    ] == expected_source_lines


def test_shipped_sample_workbook_matches_production_layout_output(tmp_path):
    """The marketing download must demonstrate the current production contract."""
    from parsers.layout_replica_parser import create_layout_replica_parser

    pdf_path = REPO_ROOT / "static" / "sample-statement.pdf"
    shipped_path = REPO_ROOT / "static" / "sample-statement.xlsx"
    generated_path = tmp_path / "sample-statement.xlsx"

    parser = create_layout_replica_parser(use_ocr=False, use_paddleocr=False)
    parser.parse(str(pdf_path), pdf_path.name)
    parser.write_excel(str(generated_path))

    shipped = load_workbook(shipped_path, data_only=False)
    generated = load_workbook(generated_path, data_only=False)
    assert shipped.sheetnames == generated.sheetnames == [
        "Exact_Copy",
        "Table_Data",
        "Full_Text",
    ]

    for sheet_name in generated.sheetnames:
        assert shipped[sheet_name].dimensions == generated[sheet_name].dimensions
        assert _sheet_cells(shipped[sheet_name]) == _sheet_cells(generated[sheet_name])

    assert all(
        cell.data_type != "f"
        for sheet in shipped.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
