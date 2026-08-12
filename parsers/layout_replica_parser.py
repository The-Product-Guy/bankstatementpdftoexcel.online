#!/usr/bin/env python3
"""Visual layout replica parser for PDF-to-Excel conversion.

This parser does not normalize bank statements. It extracts visible words with
coordinates, groups them into visual lines, and writes an Excel workbook that
approximates the PDF page layout. All cell values are written as strings.
"""
from __future__ import annotations

import logging
import os
import re
import statistics
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pdf_utils import raise_if_password_protected
from .base_parser import BaseParser

logger = logging.getLogger(__name__)


@dataclass
class LayoutWord:
    text: str
    x0: float
    x1: float
    top: float
    bottom: float
    page: int
    source: str = "pdf-text"
    confidence: Optional[float] = None
    line_hint: Optional[str] = None

    @property
    def center_y(self) -> float:
        return (self.top + self.bottom) / 2

    @property
    def height(self) -> float:
        return max(1.0, self.bottom - self.top)

    @property
    def width(self) -> float:
        return max(1.0, self.x1 - self.x0)


@dataclass
class LayoutLine:
    page: int
    index: int
    top: float
    bottom: float
    center_y: float
    words: List[LayoutWord] = field(default_factory=list)
    text: str = ""

    @property
    def height(self) -> float:
        return max(1.0, self.bottom - self.top)


@dataclass
class LayoutPage:
    page_number: int
    width: float
    height: float
    words: List[LayoutWord]
    lines: List[LayoutLine]
    source: str


@dataclass
class LayoutExtractionMetadata:
    row_count: int = 0
    col_count: int = 0
    has_data: bool = False
    extraction_method: str = "layout_replica"
    pdf_type: str = ""
    document_hint: str = "layout_replica"
    confidence: str = "good"
    message: str = ""


@dataclass
class TableColumn:
    header: str
    x0: float
    x1: float
    left: float
    right: float

    @property
    def width(self) -> float:
        return max(1.0, self.right - self.left)


@dataclass(frozen=True)
class TableReplicaRow:
    page: int
    line: int
    values: List[str]
    source: str
    schema_headers: Tuple[str, ...] = field(default_factory=tuple)
    schema_signature: Tuple[str, ...] = field(default_factory=tuple)

    def __post_init__(self) -> None:
        object.__setattr__(self, "schema_headers", tuple(self.schema_headers))
        object.__setattr__(self, "schema_signature", tuple(self.schema_signature))


class LayoutReplicaParser(BaseParser):
    """Recreate visible PDF layout in Excel without transaction semantics."""

    def __init__(
        self,
        progress_callback=None,
        dpi: int = 150,
        points_per_column: float = 6.0,
        max_columns: int = 140,
        use_ocr: bool = True,
        use_paddleocr: bool = True,
    ):
        super().__init__(progress_callback)
        self.bank_name = "Layout Replica"
        self.dpi = dpi
        self.points_per_column = points_per_column
        self.max_columns = max_columns
        self.use_ocr = use_ocr
        self.use_paddleocr = use_paddleocr
        self.pages: List[LayoutPage] = []
        self.raw_table: Optional[Dict[str, Any]] = None
        self.table_columns: List[TableColumn] = []
        self.table_rows: List[TableReplicaRow] = []
        self.table_page_summaries: List[Dict[str, Any]] = []
        self.extraction_metadata = LayoutExtractionMetadata()
        self.quality_report: Dict[str, Any] = {}
        self.source_filename = ""
        self._pdf_path = ""
        self._numeric_ocr_refinement_count = 0
        self._numeric_ocr_unresolved_count = 0

    def parse(
        self,
        pdf_path: str,
        original_filename: str,
        page_start: int = 1,
        page_end: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        self.validate_pdf_file(pdf_path)
        raise_if_password_protected(pdf_path)
        self.source_filename = original_filename
        self._pdf_path = pdf_path
        self.transactions = []
        self.pages = []
        self.raw_table = None
        self.table_columns = []
        self.table_rows = []
        self.table_page_summaries = []
        self.quality_report = {}
        self._numeric_ocr_refinement_count = 0
        self._numeric_ocr_unresolved_count = 0

        import pdfplumber

        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            if total_pages <= 0:
                raise ValueError("PDF contains no pages.")
            start, end = self._resolve_page_window(total_pages, page_start, page_end)

            for page_num in range(start, end + 1):
                page = pdf.pages[page_num - 1]
                self._emit_progress(page_num - start + 1, end - start + 1, f"Reading page {page_num}")
                words = self._extract_pdf_words(page, page_num)
                source = "pdf-text" if words else "empty"
                if self.use_ocr and self._has_sparse_word_coverage(words):
                    ocr_words = self._extract_ocr_words(
                        pdf_path,
                        page_num,
                        page.width,
                        page.height,
                    )
                    if words and ocr_words:
                        merged_words = self._merge_word_layers(words, ocr_words)
                        if len(merged_words) > len(words):
                            words = merged_words
                            source = "hybrid"
                    elif ocr_words:
                        words = ocr_words
                        source = "ocr"
                lines = self._group_words_into_lines(words, page_num)
                self.pages.append(LayoutPage(
                    page_number=page_num,
                    width=float(page.width),
                    height=float(page.height),
                    words=words,
                    lines=lines,
                    source=source,
                ))

        self._build_raw_table()
        self._build_table_replica()
        self._build_metadata()
        self.quality_report = self._build_quality_report()
        return self.transactions

    def write_excel(self, output_path: str) -> None:
        """Write a lossless layout copy plus optional table convenience views."""
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        self._write_exact_copy_sheet(wb)
        self._write_table_replica_sheet(wb)
        self._write_lines_sheet(wb)
        wb.save(output_path)

    def get_quality_report(self) -> Dict[str, Any]:
        return self.quality_report or {}

    @staticmethod
    def _resolve_page_window(total_pages: int, page_start: int, page_end: Optional[int]) -> Tuple[int, int]:
        start = max(1, page_start or 1)
        end = page_end if page_end is not None else total_pages
        end = min(total_pages, end)
        if start > end:
            raise ValueError(f"Invalid page range: start={start}, end={end}, total_pages={total_pages}")
        return start, end

    def _emit_progress(self, current: int, total: int, status: str) -> None:
        if not self.progress_callback:
            return
        try:
            self.progress_callback({
                "current_page": current,
                "total_pages": total,
                "status": status,
                "stage": "text_extract",
            })
        except Exception:
            pass

    def _extract_pdf_words(self, page, page_num: int) -> List[LayoutWord]:
        try:
            extracted = page.extract_words(
                x_tolerance=1,
                y_tolerance=3,
                keep_blank_chars=False,
                use_text_flow=False,
            )
        except TypeError:
            extracted = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
        except Exception as exc:
            logger.warning("PDF word extraction failed on page %s: %s", page_num, exc)
            return []

        words = []
        for item in extracted or []:
            text = str(item.get("text") or "").strip()
            if not text:
                continue
            try:
                words.append(LayoutWord(
                    text=text,
                    x0=float(item["x0"]),
                    x1=float(item["x1"]),
                    top=float(item["top"]),
                    bottom=float(item["bottom"]),
                    page=page_num,
                    source="pdf-text",
                ))
            except (KeyError, TypeError, ValueError):
                continue
        return words

    @staticmethod
    def _has_sparse_word_coverage(words: List[LayoutWord]) -> bool:
        """Identify text layers too small to represent a statement page.

        OCR remains off for normal text PDFs. It supplements only empty layers,
        layers with at most three words, or short fragments with fewer than
        eight words, which catches stray/hidden PDF tokens without routinely
        running two extraction engines.
        """
        if len(words) <= 3:
            return True
        visible_characters = sum(len(re.sub(r"\s+", "", word.text)) for word in words)
        return len(words) < 8 and visible_characters < 24

    @staticmethod
    def _same_visual_word(first: LayoutWord, second: LayoutWord) -> bool:
        first_text = re.sub(r"\W+", "", first.text).casefold()
        second_text = re.sub(r"\W+", "", second.text).casefold()
        likely_same_text = first_text == second_text
        if (
            not likely_same_text
            and len(first_text) == len(second_text)
            and len(first_text) >= 4
        ):
            likely_same_text = sum(
                left != right for left, right in zip(first_text, second_text)
            ) <= 1
        if not first_text or not likely_same_text:
            # Overlapping boxes alone are not enough: a stray/hidden native
            # token must never suppress different visible text found by OCR.
            return False

        x_overlap = max(0.0, min(first.x1, second.x1) - max(first.x0, second.x0))
        y_overlap = max(0.0, min(first.bottom, second.bottom) - max(first.top, second.top))
        if (
            x_overlap >= min(first.width, second.width) * 0.55
            and y_overlap >= min(first.height, second.height) * 0.55
        ):
            return True

        center_x_first = (first.x0 + first.x1) / 2
        center_x_second = (second.x0 + second.x1) / 2
        return (
            abs(first.center_y - second.center_y) <= max(first.height, second.height)
            and abs(center_x_first - center_x_second) <= max(first.width, second.width)
        )

    @classmethod
    def _merge_word_layers(
        cls,
        primary_words: List[LayoutWord],
        supplemental_words: List[LayoutWord],
    ) -> List[LayoutWord]:
        """Merge coordinate layers while keeping the primary visible token once."""
        merged = list(primary_words)
        for word in supplemental_words:
            if any(cls._same_visual_word(existing, word) for existing in merged):
                continue
            merged.append(word)
        return merged

    def _extract_ocr_words(
        self,
        pdf_path: str,
        page_num: int,
        page_width: float,
        page_height: float,
    ) -> List[LayoutWord]:
        try:
            image = self._render_page_image(pdf_path, page_num)
        except Exception as exc:
            logger.warning("Unable to render page %s for OCR: %s", page_num, exc)
            return []

        words: List[LayoutWord] = []
        if self.use_paddleocr:
            try:
                from .paddleocr_processor import PaddleOCRProcessor

                processor = PaddleOCRProcessor(use_table_structure=False)
                # Exact_Copy is a fidelity view, so keep every detected word for
                # the user to review instead of silently dropping uncertain text.
                results = processor.extract_with_coordinates(image, confidence_threshold=0.0)
                words = self._ocr_results_to_layout_words(
                    results,
                    page_num,
                    page_width,
                    page_height,
                    image.width,
                    image.height,
                    source="ocr",
                )
                if words and not self._has_sparse_word_coverage(words):
                    return words
            except Exception as exc:
                logger.warning("PaddleOCR layout extraction failed on page %s: %s", page_num, exc)

        tesseract_words = self._extract_tesseract_words(image, page_num, page_width, page_height)
        if words and tesseract_words:
            words = self._merge_word_layers(words, tesseract_words)
        elif tesseract_words:
            words = tesseract_words
        if not words:
            logger.warning("OCR layout extraction returned no words on page %s", page_num)
        return words

    def _ocr_results_to_layout_words(
        self,
        results: List[Dict[str, Any]],
        page_num: int,
        page_width: float,
        page_height: float,
        image_width: int,
        image_height: int,
        source: str,
    ) -> List[LayoutWord]:
        scale_x = page_width / max(float(image_width), 1.0)
        scale_y = page_height / max(float(image_height), 1.0)
        words: List[LayoutWord] = []
        for item in results:
            text = str(item.get("text") or "").strip()
            bbox = item.get("bbox")
            if not text or not bbox:
                continue
            try:
                x0, top, x1, bottom = [float(value) for value in bbox]
                confidence = float(item.get("confidence", 0.0))
            except (TypeError, ValueError):
                continue
            for token, token_x0, token_x1 in self._split_ocr_tokens(text, x0 * scale_x, x1 * scale_x):
                words.append(LayoutWord(
                    text=token,
                    x0=token_x0,
                    x1=token_x1,
                    top=top * scale_y,
                    bottom=bottom * scale_y,
                    page=page_num,
                    source=source,
                    confidence=confidence,
                ))
        return words

    @staticmethod
    def _split_ocr_tokens(text: str, x0: float, x1: float) -> List[Tuple[str, float, float]]:
        """Split a multi-token OCR box into per-token boxes, width allocated
        proportionally by character count. Paddle detection boxes often span
        several table cells on tight scans; kept whole, one box centered on the
        wrong column drags every token in it into that column."""
        tokens = text.split()
        if len(tokens) <= 1:
            return [(text, x0, x1)]
        total_chars = sum(len(token) for token in tokens) + (len(tokens) - 1)
        per_char = max(x1 - x0, 1.0) / total_chars
        result = []
        cursor = x0
        for token in tokens:
            token_width = len(token) * per_char
            result.append((token, cursor, cursor + token_width))
            cursor += token_width + per_char
        return result

    def _extract_tesseract_words(
        self,
        image,
        page_num: int,
        page_width: float,
        page_height: float,
    ) -> List[LayoutWord]:
        try:
            import pytesseract
        except ImportError:
            logger.warning("Tesseract OCR fallback unavailable; pytesseract is not installed.")
            return []

        if image.mode != "RGB":
            image = image.convert("RGB")

        configs = (
            "--psm 6 -c preserve_interword_spaces=1",
            "--psm 11",
        )

        def extract_candidate(candidate_image, config: str) -> List[LayoutWord]:
            try:
                data = pytesseract.image_to_data(
                    candidate_image,
                    output_type=pytesseract.Output.DICT,
                    config=config,
                )
            except Exception as exc:
                logger.warning("Tesseract OCR failed on page %s with %s: %s", page_num, config, exc)
                return []

            scale_x = page_width / max(float(candidate_image.width), 1.0)
            scale_y = page_height / max(float(candidate_image.height), 1.0)
            words: List[LayoutWord] = []
            texts = data.get("text", [])
            for idx, raw_text in enumerate(texts):
                text = str(raw_text or "").strip()
                if not text:
                    continue
                try:
                    confidence = float(data.get("conf", [0])[idx])
                except (TypeError, ValueError, IndexError):
                    confidence = 0.0
                # Tesseract uses negative confidence for non-word entries. Keep
                # all detected words with a nonnegative score so low-confidence
                # statement data remains visible in Exact_Copy for review.
                if confidence < 0:
                    continue
                try:
                    left = float(data["left"][idx])
                    top = float(data["top"][idx])
                    width = float(data["width"][idx])
                    height = float(data["height"][idx])
                except (KeyError, TypeError, ValueError, IndexError):
                    continue
                if width <= 0 or height <= 0:
                    continue
                normalized_confidence = min(confidence, 100.0) / 100.0
                words.append(LayoutWord(
                    text=text,
                    x0=left * scale_x,
                    x1=(left + width) * scale_x,
                    top=top * scale_y,
                    bottom=(top + height) * scale_y,
                    page=page_num,
                    source="ocr-tesseract",
                    confidence=normalized_confidence,
                ))
            return words

        candidates: List[Tuple[List[LayoutWord], Dict[str, float]]] = []
        for config in configs:
            words = extract_candidate(image, config)
            if words:
                candidates.append((words, self._tesseract_table_metrics(words, page_num)))

        best_original_metrics = max(
            (metrics for _words, metrics in candidates),
            key=self._tesseract_candidate_score,
            default=None,
        )
        if self._needs_preprocessed_tesseract_candidate(best_original_metrics):
            processed_image = self._preprocess_tesseract_image(image)
            if processed_image is not None:
                # Sparse-layout mode performed best on noisy scans and adds one
                # bounded OCR pass. The canvas is unchanged, so coordinates use
                # the same direct PDF scaling as the original image.
                words = extract_candidate(processed_image, "--psm 11")
                if words:
                    candidates.append((words, self._tesseract_table_metrics(words, page_num)))

            best_current_metrics = max(
                (metrics for _words, metrics in candidates),
                key=self._tesseract_candidate_score,
                default=None,
            )
            if self._needs_preprocessed_tesseract_candidate(best_current_metrics):
                table_words = self._extract_rectified_table_tesseract_words(
                    image,
                    page_num,
                    page_width,
                    page_height,
                    pytesseract,
                )
                if table_words:
                    candidates.append((
                        table_words,
                        self._tesseract_table_metrics(table_words, page_num),
                    ))

        if not candidates:
            return []
        # Dates, date-and-amount rows, and table headers outrank raw OCR volume;
        # word count and confidence only break otherwise comparable candidates.
        return max(candidates, key=lambda candidate: self._tesseract_candidate_score(candidate[1]))[0]

    @staticmethod
    def _preprocess_tesseract_image(image):
        """Denoise and Otsu-binarize without changing the coordinate canvas."""
        try:
            import cv2
            import numpy as np
            from PIL import Image

            rgb = image if image.mode == "RGB" else image.convert("RGB")
            pixels = np.asarray(rgb)
            grayscale = cv2.cvtColor(pixels, cv2.COLOR_RGB2GRAY)
            denoised = cv2.GaussianBlur(grayscale, (3, 3), 0)
            _threshold, binary = cv2.threshold(
                denoised,
                0,
                255,
                cv2.THRESH_BINARY + cv2.THRESH_OTSU,
            )
            processed = Image.fromarray(binary).convert("RGB")
            if processed.size != rgb.size:
                return None
            return processed
        except Exception as exc:
            logger.warning("Tesseract preprocessing unavailable: %s", exc)
            return None

    def _extract_rectified_table_tesseract_words(
        self,
        image,
        page_num: int,
        page_width: float,
        page_height: float,
        pytesseract,
    ) -> List[LayoutWord]:
        """OCR a strongly bordered table row-by-row and map boxes to the page.

        This bounded fallback runs only after ordinary OCR lacks table coverage.
        It requires one dominant four-corner table and a reliable horizontal
        grid, so borderless documents continue through the cheaper candidates.
        """
        try:
            import cv2
            import numpy as np

            rgb_image = image if image.mode == "RGB" else image.convert("RGB")
            pixels = np.asarray(rgb_image)
            grayscale = cv2.cvtColor(pixels, cv2.COLOR_RGB2GRAY)
            blurred = cv2.GaussianBlur(grayscale, (3, 3), 0)
            _threshold, inverted = cv2.threshold(
                blurred,
                0,
                255,
                cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU,
            )
            closed = cv2.morphologyEx(
                inverted,
                cv2.MORPH_CLOSE,
                np.ones((5, 5), dtype=np.uint8),
            )
            contours = cv2.findContours(
                closed,
                cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE,
            )[0]
            if not contours:
                return []
            contour = max(contours, key=cv2.contourArea)
            image_area = max(float(pixels.shape[0] * pixels.shape[1]), 1.0)
            contour_ratio = cv2.contourArea(contour) / image_area
            if not 0.25 <= contour_ratio <= 0.9:
                return []
            perimeter = cv2.arcLength(contour, True)
            quad = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
            if len(quad) != 4 or not cv2.isContourConvex(quad):
                return []

            points = quad.reshape(4, 2).astype(np.float32)
            sums = points.sum(axis=1)
            differences = np.diff(points, axis=1).reshape(-1)
            ordered = np.zeros((4, 2), dtype=np.float32)
            ordered[0] = points[np.argmin(sums)]  # top-left
            ordered[1] = points[np.argmin(differences)]  # top-right
            ordered[2] = points[np.argmax(sums)]  # bottom-right
            ordered[3] = points[np.argmax(differences)]  # bottom-left
            top_left, top_right, bottom_right, bottom_left = ordered
            target_width = int(max(
                np.linalg.norm(bottom_right - bottom_left),
                np.linalg.norm(top_right - top_left),
            ))
            target_height = int(max(
                np.linalg.norm(top_right - bottom_right),
                np.linalg.norm(top_left - bottom_left),
            ))
            if target_width < 300 or target_height < 300:
                return []

            destination = np.array([
                [0, 0],
                [target_width - 1, 0],
                [target_width - 1, target_height - 1],
                [0, target_height - 1],
            ], dtype=np.float32)
            transform = cv2.getPerspectiveTransform(ordered, destination)
            inverse_transform = cv2.getPerspectiveTransform(destination, ordered)
            rectified = cv2.warpPerspective(
                pixels,
                transform,
                (target_width, target_height),
                borderValue=(255, 255, 255),
            )
            rectified_gray = cv2.cvtColor(rectified, cv2.COLOR_RGB2GRAY)
            rectified_blur = cv2.GaussianBlur(rectified_gray, (3, 3), 0)
            _threshold, rectified_binary = cv2.threshold(
                rectified_blur,
                0,
                255,
                cv2.THRESH_BINARY + cv2.THRESH_OTSU,
            )
            rectified_inverted = 255 - rectified_binary
            horizontal_mask = cv2.morphologyEx(
                rectified_inverted,
                cv2.MORPH_OPEN,
                cv2.getStructuringElement(
                    cv2.MORPH_RECT,
                    (max(20, target_width // 8), 1),
                ),
            )
            projection = np.count_nonzero(horizontal_mask, axis=1)
            boundary_pixels = np.where(projection > target_width * 0.35)[0]
            boundary_groups: List[List[int]] = []
            for raw_y in boundary_pixels:
                y = int(raw_y)
                if not boundary_groups or y > boundary_groups[-1][-1] + 1:
                    boundary_groups.append([y])
                else:
                    boundary_groups[-1].append(y)
            boundaries = [
                int(round(sum(group) / len(group)))
                for group in boundary_groups
            ]
            # Thick/noisy top borders can appear as several lines within the
            # first 3% of the table. Keep the outermost one and discard echoes.
            later_boundaries = [
                boundary for boundary in boundaries
                if boundary >= target_height * 0.03
            ]
            if boundaries and later_boundaries:
                boundaries = [boundaries[0], *later_boundaries]
            boundaries = list(dict.fromkeys(boundaries))
            bands = [
                (top, bottom)
                for top, bottom in zip(boundaries, boundaries[1:])
                if bottom - top >= 18
            ]
            if not 6 <= len(bands) <= 64:
                return []

            scale_x = page_width / max(float(rgb_image.width), 1.0)
            scale_y = page_height / max(float(rgb_image.height), 1.0)
            extracted_words: List[LayoutWord] = []
            header_vocabulary = {
                "date", "description", "amount", "balance", "debit", "debits",
                "credit", "credits", "withdrawal", "withdrawals", "deposit", "deposits",
            }

            for band_index, (band_top, band_bottom) in enumerate(bands):
                crop_x = 2
                crop_y = band_top + 2
                crop = rectified_binary[
                    crop_y:max(crop_y + 1, band_bottom - 2),
                    crop_x:max(crop_x + 1, target_width - 2),
                ]
                if crop.size == 0:
                    continue
                primary_mode = "--psm 7" if crop.shape[0] < 100 else "--psm 6"
                modes = [
                    primary_mode,
                    *[
                        mode for mode in ("--psm 11", "--psm 6", "--psm 7")
                        if mode != primary_mode
                    ],
                ]
                best_band_words: List[LayoutWord] = []
                best_band_score: Tuple[float, ...] = (-1.0,)
                for mode in modes:
                    try:
                        data = pytesseract.image_to_data(
                            crop,
                            output_type=pytesseract.Output.DICT,
                            config=mode,
                        )
                    except Exception:
                        continue
                    band_words: List[LayoutWord] = []
                    texts = data.get("text", [])
                    for index, raw_text in enumerate(texts):
                        text = str(raw_text or "").strip()
                        if not text:
                            continue
                        try:
                            confidence = float(data.get("conf", [0])[index])
                            left = float(data["left"][index]) + crop_x
                            top = float(data["top"][index]) + crop_y
                            width = float(data["width"][index])
                            height = float(data["height"][index])
                        except (KeyError, TypeError, ValueError, IndexError):
                            continue
                        if confidence < 0 or width <= 0 or height <= 0:
                            continue
                        corners = np.array([[[
                            left,
                            top,
                        ], [
                            left + width,
                            top,
                        ], [
                            left + width,
                            top + height,
                        ], [
                            left,
                            top + height,
                        ]]], dtype=np.float32)
                        mapped = cv2.perspectiveTransform(corners, inverse_transform)[0]
                        x_values = np.clip(mapped[:, 0], 0, rgb_image.width)
                        y_values = np.clip(mapped[:, 1], 0, rgb_image.height)

                        def data_number(key: str) -> int:
                            try:
                                return int(data.get(key, [0])[index])
                            except (TypeError, ValueError, IndexError):
                                return 0

                        line_hint = (
                            f"table-{page_num}-{band_index}-"
                            f"{data_number('block_num')}-{data_number('par_num')}-"
                            f"{data_number('line_num')}"
                        )
                        band_words.append(LayoutWord(
                            text=text,
                            x0=float(min(x_values)) * scale_x,
                            x1=float(max(x_values)) * scale_x,
                            top=float(min(y_values)) * scale_y,
                            bottom=float(max(y_values)) * scale_y,
                            page=page_num,
                            source="ocr-tesseract-table",
                            confidence=min(confidence, 100.0) / 100.0,
                            line_hint=line_hint,
                        ))
                    date_hits = sum(self._looks_like_date_value(word.text) for word in band_words)
                    header_hits = sum(
                        word.text.casefold().strip(".:") in header_vocabulary
                        for word in band_words
                    )
                    confidences = [word.confidence or 0.0 for word in band_words]
                    band_score = (
                        float(date_hits > 0),
                        float(header_hits),
                        float(len(band_words)),
                        sum(confidences) / len(confidences) if confidences else 0.0,
                    )
                    if band_score > best_band_score:
                        best_band_score = band_score
                        best_band_words = band_words
                    # A date row, a recognizable header, or a populated band is
                    # sufficient; only weak/empty bands pay for another mode.
                    if date_hits or header_hits >= 2 or len(band_words) >= 3:
                        break
                extracted_words.extend(best_band_words)
            return extracted_words
        except Exception as exc:
            logger.warning("Rectified table OCR unavailable on page %s: %s", page_num, exc)
            return []

    def _tesseract_table_metrics(
        self,
        words: List[LayoutWord],
        page_num: int,
    ) -> Dict[str, float]:
        header_vocabulary = {
            "date", "description", "amount", "balance", "debit", "credit",
            "withdrawal", "withdrawals", "deposit", "deposits", "reference",
            "particulars", "narration", "details", "value", "money", "out", "in",
        }
        date_count = sum(self._looks_like_date_value(word.text) for word in words)
        amount_count = sum(self._looks_like_amount_value(word.text) for word in words)
        header_count = sum(
            word.text.casefold().strip(".:") in header_vocabulary
            for word in words
        )
        transaction_line_count = 0
        for line in self._group_words_into_lines(words, page_num):
            has_date = any(self._looks_like_date_value(word.text) for word in line.words)
            has_amount = any(self._looks_like_amount_value(word.text) for word in line.words)
            if has_date and has_amount:
                transaction_line_count += 1
        confidences = [word.confidence for word in words if word.confidence is not None]
        average_confidence = sum(confidences) / len(confidences) if confidences else 0.0
        return {
            "transaction_lines": float(transaction_line_count),
            "dates": float(date_count),
            "amounts": float(amount_count),
            "headers": float(header_count),
            "words": float(len(words)),
            "confidence": average_confidence,
        }

    @staticmethod
    def _tesseract_candidate_score(metrics: Dict[str, float]) -> Tuple[float, ...]:
        dates = metrics["dates"]
        amounts = metrics["amounts"]
        table_score = (
            (30.0 * metrics["transaction_lines"])
            + (10.0 * dates)
            + (4.0 * metrics["headers"])
            + (2.0 * min(amounts, max(4.0, dates * 4.0)))
        )
        return (
            table_score,
            dates,
            metrics["transaction_lines"],
            metrics["headers"],
            metrics["amounts"],
            metrics["words"],
            metrics["confidence"],
        )

    @staticmethod
    def _needs_preprocessed_tesseract_candidate(
        metrics: Optional[Dict[str, float]],
    ) -> bool:
        if not metrics:
            return True
        # One extra pass is reserved for candidates that do not yet resemble a
        # statement table. Normal pages with solid row coverage keep two passes.
        return metrics["dates"] < 8 or metrics["transaction_lines"] < 6

    def _render_page_image(
        self,
        pdf_path: str,
        page_num: int,
        dpi: Optional[int] = None,
    ):
        render_dpi = int(dpi or self.dpi)
        try:
            import pymupdf
            from PIL import Image

            doc = pymupdf.open(pdf_path)
            try:
                page = doc.load_page(page_num - 1)
                scale = render_dpi / 72.0
                pix = page.get_pixmap(matrix=pymupdf.Matrix(scale, scale), alpha=False)
                return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            finally:
                doc.close()
        except ImportError:
            from pdf2image import convert_from_path

            images = convert_from_path(
                pdf_path,
                dpi=render_dpi,
                first_page=page_num,
                last_page=page_num,
            )
            if not images:
                raise RuntimeError("Unable to render PDF page for OCR.")
            return images[0].convert("RGB")

    def _group_words_into_lines(self, words: List[LayoutWord], page_num: int) -> List[LayoutLine]:
        if not words:
            return []

        heights = [word.height for word in words]
        median_height = statistics.median(heights) if heights else 8.0
        line_tolerance = max(2.5, median_height * 0.45)
        sorted_words = sorted(words, key=lambda word: (word.center_y, word.x0))
        hinted_buckets: Dict[str, List[LayoutWord]] = {}
        unhinted_words: List[LayoutWord] = []
        for word in sorted_words:
            if word.line_hint:
                hinted_buckets.setdefault(word.line_hint, []).append(word)
            else:
                unhinted_words.append(word)

        line_buckets = list(hinted_buckets.values())
        centers = [
            sum(word.center_y for word in bucket) / len(bucket)
            for bucket in line_buckets
        ]

        for word in unhinted_words:
            matched_index = None
            for idx, center in enumerate(centers):
                if abs(word.center_y - center) <= line_tolerance:
                    matched_index = idx
                    break
            if matched_index is None:
                line_buckets.append([word])
                centers.append(word.center_y)
            else:
                line_buckets[matched_index].append(word)
                bucket = line_buckets[matched_index]
                centers[matched_index] = sum(item.center_y for item in bucket) / len(bucket)

        lines: List[LayoutLine] = []
        for idx, bucket in enumerate(sorted(line_buckets, key=lambda group: min(word.top for word in group)), 1):
            bucket.sort(key=lambda word: word.x0)
            top = min(word.top for word in bucket)
            bottom = max(word.bottom for word in bucket)
            center_y = sum(word.center_y for word in bucket) / len(bucket)
            lines.append(LayoutLine(
                page=page_num,
                index=idx,
                top=top,
                bottom=bottom,
                center_y=center_y,
                words=bucket,
                text=self._line_text(bucket),
            ))
        return lines

    def _line_text(self, words: List[LayoutWord]) -> str:
        if not words:
            return ""
        char_widths = [word.width / max(len(word.text), 1) for word in words if word.text]
        char_width = statistics.median(char_widths) if char_widths else self.points_per_column
        char_width = max(2.0, min(char_width, 9.0))

        leading_spaces = " " * max(0, min(24, round(words[0].x0 / char_width)))
        parts = [f"{leading_spaces}{words[0].text}"]
        previous = words[0]
        for word in words[1:]:
            gap = word.x0 - previous.x1
            if gap <= 0.75:
                spaces = ""
            else:
                spaces = " " * max(1, min(16, round(gap / char_width)))
            parts.append(f"{spaces}{word.text}")
            previous = word
        return "".join(parts)

    def _column_for_x(self, x0: float, page_width: float) -> int:
        col = int(round(max(0.0, x0) / self.points_per_column)) + 1
        page_max = int(round(page_width / self.points_per_column)) + 2
        return max(1, min(col, min(self.max_columns, page_max)))

    def _write_page_sheet_rows(
        self,
        sheet,
        page: LayoutPage,
        start_row: int,
        include_page_header: bool,
    ) -> int:
        max_col = min(self.max_columns, int(round(page.width / self.points_per_column)) + 2)
        self._prepare_replica_sheet(sheet, max_col)
        row_idx = start_row

        if include_page_header:
            cell = sheet.cell(row=row_idx, column=1)
            self._set_text_cell(cell, f"Page {page.page_number}")
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2F8")
            row_idx += 1

        for line in page.lines:
            height = max(12, min(28, line.height * 1.25))
            sheet.row_dimensions[row_idx].height = height
            words_by_column: Dict[int, List[str]] = {}
            for word in line.words:
                col_idx = self._column_for_x(word.x0, page.width)
                words_by_column.setdefault(col_idx, []).append(str(word.text))

            for col_idx, word_texts in words_by_column.items():
                cell = sheet.cell(row=row_idx, column=col_idx)
                self._set_text_cell(cell, " ".join(word_texts))
                cell.alignment = Alignment(vertical="top", wrap_text=False)
            row_idx += 1
        return row_idx

    def _write_exact_copy_sheet(self, wb: Workbook) -> None:
        """Write every extracted visual line once, retaining page geometry."""
        sheet = wb.create_sheet("Exact_Copy")
        row_idx = 1
        for page in self.pages:
            row_idx = self._write_page_sheet_rows(
                sheet,
                page,
                start_row=row_idx,
                include_page_header=True,
            )

    @staticmethod
    def _set_text_cell(cell, value: Any) -> None:
        """Store a value as literal text, including strings beginning with '='."""
        text = "" if value is None else str(value)
        cell.value = text
        cell.data_type = "s"
        cell.number_format = "@"
        if text.startswith("="):
            cell.quotePrefix = True

    def _prepare_replica_sheet(self, sheet, max_col: int) -> None:
        sheet.sheet_view.showGridLines = True
        for col_idx in range(1, max_col + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 1.6

    def _build_table_replica(self) -> None:
        """Build a table-first view using inferred table headers and columns."""
        active_columns: List[TableColumn] = []
        table_columns_status = ""

        def adopt_workbook_columns(columns: List[TableColumn], status: str) -> None:
            # Real header rows always outrank positional inference; a single
            # noisy page must not replace detected bank headers just because
            # its word clusters produced more columns.
            nonlocal table_columns_status
            if table_columns_status == "header" and status != "header":
                return
            if status != "header" or table_columns_status == "header":
                if len(columns) <= len(self.table_columns):
                    return
            self.table_columns = [
                TableColumn(col.header, col.x0, col.x1, col.left, col.right)
                for col in columns
            ]
            table_columns_status = status

        for page in self.pages:
            header_line = self._detect_table_header_line(page)
            if (
                header_line
                and active_columns
                and self._header_vocabulary_hits(header_line) < 2
            ):
                # Continuation pages can contain an uppercase transaction row
                # that scores like a header. Keep carried columns unless the
                # new page contains at least two real header labels.
                header_line = None
            page_columns: List[TableColumn] = []
            column_status = "carried"

            if header_line:
                page_columns = self._columns_from_header_line(header_line, page.width)
                page_columns = self._prepend_missing_date_column(page, page_columns)
                if len(page_columns) >= 2:
                    active_columns = page_columns
                    column_status = "header"
                    adopt_workbook_columns(page_columns, "header")

            if not active_columns:
                inferred_columns = self._columns_from_transaction_lines(page)
                if inferred_columns:
                    active_columns = inferred_columns
                    column_status = "inferred"
                    adopt_workbook_columns(inferred_columns, "inferred")
                else:
                    self.table_page_summaries.append({
                        "page": page.page_number,
                        "header_line": "",
                        "columns": 0,
                        "rows": 0,
                        "status": "no_table_header",
                    })
                    continue

            before_count = len(self.table_rows)
            page_rows = self._table_rows_from_page(page, active_columns, header_line)

            if not page_rows:
                inferred_columns = self._columns_from_transaction_lines(page)
                if inferred_columns:
                    inferred_rows = self._table_rows_from_page(page, inferred_columns, None)
                    if inferred_rows:
                        active_columns = inferred_columns
                        column_status = "inferred"
                        header_line = None
                        page_rows = inferred_rows
                        adopt_workbook_columns(inferred_columns, "inferred")

            self.table_rows.extend(page_rows)

            self.table_page_summaries.append({
                "page": page.page_number,
                "header_line": str(header_line.index) if header_line and column_status == "header" else column_status,
                "columns": len(active_columns),
                "rows": len(self.table_rows) - before_count,
                "status": "ok" if len(self.table_rows) > before_count else "no_table_rows",
            })

        if not self.table_columns and self.table_rows:
            max_width = max(len(row.values) for row in self.table_rows)
            self.table_columns = [
                TableColumn(f"Column {idx}", 0.0, 0.0, 0.0, 0.0)
                for idx in range(1, max_width + 1)
            ]

    def _table_rows_from_page(
        self,
        page: LayoutPage,
        columns: List[TableColumn],
        header_line: Optional[LayoutLine],
    ) -> List[TableReplicaRow]:
        anchored_rows = self._table_rows_from_ocr_date_anchors(
            page,
            columns,
            header_line,
        )
        if anchored_rows is not None:
            return anchored_rows

        rows: List[TableReplicaRow] = []
        saw_data_row = False
        for line in page.lines:
            if self._is_separator_line(line):
                continue
            if header_line and line.index == header_line.index:
                continue

            values = self._assign_line_to_table_columns(line, columns)
            populated = sum(1 for value in values if value)
            if not populated:
                continue

            # Wrapped cell text belongs to the previous row, not a new one.
            if self._is_continuation_row(line, values, columns, rows):
                self._merge_continuation_row(rows[-1], values)
                continue

            if not self._is_table_relevant_row(line, values, columns, saw_data_row):
                continue

            if populated == 1 and not saw_data_row:
                continue
            if populated >= 2:
                saw_data_row = True

            rows.append(TableReplicaRow(
                page=page.page_number,
                line=line.index,
                values=values,
                source=page.source,
                schema_headers=tuple(column.header for column in columns),
                schema_signature=self._table_schema_signature(columns),
            ))
        return rows

    def _table_rows_from_ocr_date_anchors(
        self,
        page: LayoutPage,
        columns: List[TableColumn],
        header_line: Optional[LayoutLine],
    ) -> Optional[List[TableReplicaRow]]:
        """Rebuild skewed OCR table rows around unambiguous date anchors.

        This path is intentionally conservative.  It is used only when an OCR
        page has an explicit Date column and exactly one rightmost numeric value
        for every detected transaction date.  The right anchor can be Balance,
        or the sole Amount column in a compact card statement.  Those two
        monotonic streams let us measure the page's vertical perspective offset
        and associate split description/amount fragments with the correct date
        without changing the source lines used by ``Exact_Copy``.
        """
        if page.source not in {"ocr", "hybrid"}:
            return None

        date_indices = [
            idx
            for idx, column in enumerate(columns)
            if "date" in column.header.casefold()
        ]
        balance_indices = [
            idx
            for idx, column in enumerate(columns)
            if "balance" in column.header.casefold()
            or "saldo" in self._header_token_base(column.header)
        ]
        generic_amount_indices = [
            idx
            for idx, column in enumerate(columns)
            if self._header_token_base(column.header) in {"amount", "betrag", "importe"}
        ]
        if len(date_indices) != 1:
            return None
        if len(balance_indices) == 1:
            right_anchor_idx = balance_indices[0]
            has_balance_anchor = True
        elif not balance_indices and len(generic_amount_indices) == 1:
            right_anchor_idx = generic_amount_indices[0]
            has_balance_anchor = False
        else:
            return None

        header_word_ids = {
            id(word)
            for word in (header_line.words if header_line else [])
        }
        date_words = sorted(
            (
                word
                for word in page.words
                if id(word) not in header_word_ids
                and self._looks_like_date_value(word.text)
            ),
            key=lambda word: word.center_y,
        )
        if len(date_words) < 4:
            return None

        right_anchor_column = columns[right_anchor_idx]
        right_header_center = (right_anchor_column.x0 + right_anchor_column.x1) / 2.0
        right_anchor_threshold = (
            right_header_center
            if has_balance_anchor
            else right_anchor_column.x0 - max(6.0, right_anchor_column.width * 0.08)
        )
        right_anchor_words = sorted(
            (
                word
                for word in page.words
                if id(word) not in header_word_ids
                and self._looks_like_amount_value(word.text)
                and (word.x0 + word.x1) / 2.0 >= right_anchor_threshold
            ),
            key=lambda word: word.center_y,
        )
        if len(right_anchor_words) != len(date_words):
            return None

        date_gaps = [
            current.center_y - previous.center_y
            for previous, current in zip(date_words, date_words[1:])
            if current.center_y - previous.center_y > 8.0
        ]
        if not date_gaps:
            return None
        row_gap = statistics.median(date_gaps)
        right_anchor_offsets = [
            right_anchor.center_y - date.center_y
            for date, right_anchor in zip(date_words, right_anchor_words)
        ]
        vertical_offset = statistics.median(right_anchor_offsets)
        offset_deviations = sorted(
            abs(offset - vertical_offset)
            for offset in right_anchor_offsets
        )
        percentile_80 = offset_deviations[
            min(len(offset_deviations) - 1, int(len(offset_deviations) * 0.8))
        ]
        if percentile_80 > max(4.0, row_gap * 0.22):
            return None

        date_x = statistics.median(
            (word.x0 + word.x1) / 2.0
            for word in date_words
        )
        right_anchor_x = statistics.median(
            (word.x0 + word.x1) / 2.0
            for word in right_anchor_words
        )
        horizontal_span = right_anchor_x - date_x
        if horizontal_span <= 40.0:
            return None

        excluded_word_ids = set(header_word_ids)
        line_for_word: Dict[int, int] = {}
        for line in page.lines:
            for word in line.words:
                line_for_word[id(word)] = line.index
            is_repeated_header = (
                self._header_vocabulary_hits(line) >= 2
                and not any(
                    self._looks_like_date_value(word.text)
                    or self._looks_like_amount_value(word.text)
                    for word in line.words
                )
            )
            if (
                self._is_separator_line(line)
                or self._is_document_context_line(line.text)
                or is_repeated_header
            ):
                excluded_word_ids.update(id(word) for word in line.words)

        buckets: List[List[List[LayoutWord]]] = [
            [[] for _column in columns]
            for _date in date_words
        ]
        amount_cluster_split = None
        if has_balance_anchor:
            amount_cluster_split = self._ocr_transaction_amount_cluster_split(
                page,
                columns,
                date_x,
                right_anchor_x,
                header_word_ids,
            )
        date_row_by_word_id = {
            id(word): idx
            for idx, word in enumerate(date_words)
        }
        maximum_residual = max(8.0, row_gap * 0.72)

        for word in page.words:
            if id(word) in excluded_word_ids:
                continue
            column_idx = self._anchored_table_column_index_for_word(
                word,
                columns,
                amount_cluster_split,
            )
            if column_idx is None:
                continue

            row_idx = date_row_by_word_id.get(id(word))
            if row_idx is None:
                word_x = (word.x0 + word.x1) / 2.0
                x_ratio = (word_x - date_x) / horizontal_span
                x_ratio = max(-0.15, min(1.20, x_ratio))
                expected_offset = vertical_offset * x_ratio
                residuals = [
                    abs(word.center_y - (date.center_y + expected_offset))
                    for date in date_words
                ]
                row_idx = min(range(len(residuals)), key=residuals.__getitem__)
                if residuals[row_idx] > maximum_residual:
                    continue

            buckets[row_idx][column_idx].append(word)

        schema_headers = tuple(column.header for column in columns)
        schema_signature = self._table_schema_signature(columns)
        values_by_row: List[List[str]] = []
        for date_word, column_words in zip(date_words, buckets):
            values = [
                self._join_ocr_anchor_words(
                    words,
                    vertical_offset,
                    date_x,
                    horizontal_span,
                    row_gap,
                )
                for words in column_words
            ]
            if not values[date_indices[0]]:
                values[date_indices[0]] = date_word.text
            values_by_row.append(values)

        # Exact_Copy remains a verbatim view of ``page.words``.  Table_Data may
        # use a second, cell-local OCR reading when the first pass produced an
        # invalid money token or a ledger inconsistency.  The refinement never
        # changes source words and only accepts two agreeing direct OCR reads.
        self._refine_ocr_numeric_cells(page, columns, buckets, values_by_row)

        rows: List[TableReplicaRow] = []
        for row_idx, (date_word, values) in enumerate(zip(date_words, values_by_row)):
            rows.append(TableReplicaRow(
                page=page.page_number,
                line=line_for_word.get(id(date_word), row_idx + 1),
                values=values,
                source=page.source,
                schema_headers=schema_headers,
                schema_signature=schema_signature,
            ))
        return rows

    def _anchored_table_column_index_for_word(
        self,
        word: LayoutWord,
        columns: List[TableColumn],
        amount_cluster_split: Optional[Tuple[float, int, int, float, float]] = None,
    ) -> Optional[int]:
        date_idx = next(
            (
                idx
                for idx, column in enumerate(columns)
                if "date" in column.header.casefold()
            ),
            None,
        )
        if date_idx is not None and self._looks_like_date_value(word.text):
            return date_idx

        column_idx = self._table_column_index_for_word(word, columns)
        if column_idx is None:
            return None

        center_x = (word.x0 + word.x1) / 2.0
        if self._looks_like_amount_value(word.text):
            if amount_cluster_split is not None:
                (
                    split_x,
                    left_column_idx,
                    right_column_idx,
                    transaction_band_left,
                    transaction_band_right,
                ) = amount_cluster_split
                if transaction_band_left <= center_x < transaction_band_right:
                    return left_column_idx if center_x < split_x else right_column_idx
            if (
                column_idx > 0
                and self._is_amount_header(columns[column_idx].header)
                and self._is_amount_header(columns[column_idx - 1].header)
                and center_x < columns[column_idx].x0
            ):
                # Wide OCR boxes can straddle an amount-column boundary.  The
                # next header's x anchor is stronger evidence than a 1-2 point
                # overlap advantage (Debit vs Credit, or Credit vs Balance).
                return column_idx - 1
            return column_idx

        text_column_limit: Optional[float] = None
        if amount_cluster_split is not None:
            (
                _split_x,
                left_column_idx,
                right_column_idx,
                transaction_band_left,
                _transaction_band_right,
            ) = amount_cluster_split
            if column_idx in {left_column_idx, right_column_idx}:
                text_column_limit = transaction_band_left
        if self._is_amount_header(columns[column_idx].header) and (
            center_x < columns[column_idx].x0
            or (text_column_limit is not None and center_x < text_column_limit)
        ):
            for previous_idx in range(column_idx - 1, -1, -1):
                if self._is_wide_text_header(columns[previous_idx].header):
                    return previous_idx
        return column_idx

    def _ocr_transaction_amount_cluster_split(
        self,
        page: LayoutPage,
        columns: List[TableColumn],
        observed_date_x: float,
        observed_balance_x: float,
        header_word_ids: set[int],
    ) -> Optional[Tuple[float, int, int, float, float]]:
        """Map Debit/Credit anchors into the current page's observed geometry.

        Date and Balance provide a stable affine baseline even when a page has
        only one transaction side (or a single minority-side value).  Two
        well-supported amount clusters may refine the midpoint, but are never
        required for calibration.
        """
        date_idx = next(
            (
                idx
                for idx, column in enumerate(columns)
                if "date" in column.header.casefold()
            ),
            None,
        )
        debit_idx = next(
            (
                idx
                for idx, column in enumerate(columns)
                if any(
                    token in column.header.casefold()
                    for token in ("debit", "withdraw", "money out", "soll", "cargo")
                )
            ),
            None,
        )
        credit_idx = next(
            (
                idx
                for idx, column in enumerate(columns)
                if any(
                    token in column.header.casefold()
                    for token in ("credit", "deposit", "money in", "haben", "abono")
                )
            ),
            None,
        )
        balance_idx = next(
            (
                idx
                for idx, column in enumerate(columns)
                if "balance" in column.header.casefold()
                or "saldo" in self._header_token_base(column.header)
            ),
            None,
        )
        if (
            date_idx is None
            or debit_idx is None
            or credit_idx is None
            or balance_idx is None
            or not (date_idx < debit_idx < credit_idx < balance_idx)
        ):
            return None

        header_centers = [
            (column.x0 + column.x1) / 2.0
            for column in columns
        ]
        header_span = header_centers[balance_idx] - header_centers[date_idx]
        observed_span = observed_balance_x - observed_date_x
        if header_span <= 40.0 or observed_span <= 40.0:
            return None
        scale = observed_span / header_span
        if not 0.50 <= scale <= 1.80:
            return None

        def mapped_anchor(column_idx: int) -> float:
            return observed_date_x + (
                (header_centers[column_idx] - header_centers[date_idx]) * scale
            )

        mapped_debit_x = mapped_anchor(debit_idx)
        mapped_credit_x = mapped_anchor(credit_idx)
        mapped_balance_x = mapped_anchor(balance_idx)
        amount_anchor_gap = mapped_credit_x - mapped_debit_x
        if amount_anchor_gap < 25.0 or mapped_balance_x <= mapped_credit_x:
            return None

        geometric_split = (mapped_debit_x + mapped_credit_x) / 2.0
        transaction_band_left = mapped_debit_x - (amount_anchor_gap * 0.55)
        transaction_band_right = (mapped_credit_x + mapped_balance_x) / 2.0

        amount_words = sorted(
            (
                (word.x0 + word.x1) / 2.0,
                word.width,
            )
            for word in page.words
            if id(word) not in header_word_ids
            and self._looks_like_amount_value(word.text)
            and transaction_band_left
            <= (word.x0 + word.x1) / 2.0
            < transaction_band_right
        )
        centers = [center for center, _width in amount_words]
        split_x = geometric_split
        if len(centers) >= 4:
            gaps = [
                (right - left, idx)
                for idx, (left, right) in enumerate(zip(centers, centers[1:]))
            ]
            largest_gap, split_idx = max(gaps)
            left_centers = centers[:split_idx + 1]
            right_centers = centers[split_idx + 1:]
            cluster_split = (left_centers[-1] + right_centers[0]) / 2.0
            if (
                largest_gap >= max(30.0, amount_anchor_gap * 0.30)
                and len(left_centers) >= 2
                and len(right_centers) >= 2
                and abs(cluster_split - geometric_split) <= amount_anchor_gap * 0.55
            ):
                split_x = cluster_split

        return (
            split_x,
            debit_idx,
            credit_idx,
            transaction_band_left,
            transaction_band_right,
        )

    @staticmethod
    def _join_ocr_anchor_words(
        words: List[LayoutWord],
        vertical_offset: float,
        date_x: float,
        horizontal_span: float,
        row_gap: float,
    ) -> str:
        if not words:
            return ""
        band_height = max(4.0, min(12.0, row_gap * 0.30))

        def corrected_y(word: LayoutWord) -> float:
            center_x = (word.x0 + word.x1) / 2.0
            x_ratio = (center_x - date_x) / horizontal_span
            x_ratio = max(-0.15, min(1.20, x_ratio))
            return word.center_y - (vertical_offset * x_ratio)

        bands: List[Tuple[float, List[LayoutWord]]] = []
        for word in sorted(words, key=lambda item: (corrected_y(item), item.x0)):
            word_y = corrected_y(word)
            if bands and abs(word_y - bands[-1][0]) <= band_height:
                center, band_words = bands[-1]
                band_words.append(word)
                bands[-1] = (
                    ((center * (len(band_words) - 1)) + word_y) / len(band_words),
                    band_words,
                )
            else:
                bands.append((word_y, [word]))

        ordered_words = [
            word
            for _center, band_words in bands
            for word in sorted(band_words, key=lambda item: item.x0)
        ]
        return " ".join(word.text for word in ordered_words).strip()

    @staticmethod
    def _money_digit_fingerprint(text: str) -> str:
        """Digits printed in a money token, independent of sign/punctuation."""
        return "".join(re.findall(r"\d", str(text or "")))

    @staticmethod
    def _has_negative_money_marker(text: str) -> bool:
        value = re.sub(r"\s+", "", str(text or "")).upper()
        return value.startswith("-") or (
            value.startswith("(") and value.endswith(")")
        ) or value.endswith("DR")

    @classmethod
    def _preserve_negative_money_marker(cls, original: str, candidate: str) -> str:
        """Carry an already-extracted negative marker across punctuation OCR."""
        if not cls._has_negative_money_marker(original) or cls._has_negative_money_marker(candidate):
            return candidate
        source = re.sub(r"\s+", "", str(original or "")).upper()
        value = str(candidate or "").lstrip("+")
        if source.startswith("(") and source.endswith(")"):
            return f"({value})"
        if source.endswith("DR"):
            return f"{value}DR"
        return f"-{value}"

    @classmethod
    def _strict_money_value(cls, text: str) -> Optional[Decimal]:
        """Parse an unambiguous printed money token without repairing it.

        This intentionally rejects OCR strings such as ``1.498.03`` and
        ``10,889,97``.  They retain all digits but use one separator as both a
        thousands and decimal mark, so a second OCR read is required before
        Table_Data may change them.
        """
        value = re.sub(r"\s+", "", str(text or "")).strip()
        if not value:
            return None

        negative = False
        if value.startswith("(") and value.endswith(")"):
            negative = True
            value = value[1:-1]

        suffix_match = re.search(r"(CR|DR)$", value, re.IGNORECASE)
        if suffix_match:
            negative = negative or suffix_match.group(1).upper() == "DR"
            value = value[:suffix_match.start()]

        if value[:1] in "+-":
            negative = negative or value[0] == "-"
            value = value[1:]
        if value[:1] in "$£€₹":
            value = value[1:]
        if value[:1] in "+-":
            negative = negative or value[0] == "-"
            value = value[1:]
        if not value or not re.fullmatch(r"[0-9.,]+", value):
            return None

        decimal_separator: Optional[str] = None
        grouping_separator: Optional[str] = None
        if "," in value and "." in value:
            decimal_separator = "," if value.rfind(",") > value.rfind(".") else "."
            grouping_separator = "." if decimal_separator == "," else ","
        elif "," in value or "." in value:
            separator = "," if "," in value else "."
            parts = value.split(separator)
            if len(parts) == 2 and len(parts[1]) == 2:
                decimal_separator = separator
            elif len(parts) >= 2 and all(len(part) == 3 for part in parts[1:]):
                grouping_separator = separator
            else:
                return None

        integer_part = value
        fraction = ""
        if decimal_separator:
            integer_part, fraction = value.rsplit(decimal_separator, 1)
            if len(fraction) != 2 or not fraction.isdigit():
                return None

        if grouping_separator:
            groups = integer_part.split(grouping_separator)
            if not groups or not (1 <= len(groups[0]) <= 3):
                return None
            # Western grouping (1,234,567) and Indian grouping
            # (1,23,456) are both valid source formats.
            western = all(len(group) == 3 for group in groups[1:])
            indian = (
                len(groups) >= 2
                and len(groups[-1]) == 3
                and all(len(group) == 2 for group in groups[1:-1])
            )
            if not (western or indian) or not all(group.isdigit() for group in groups):
                return None
            integer_digits = "".join(groups)
        else:
            if not integer_part.isdigit():
                return None
            integer_digits = integer_part

        normalized = integer_digits + (f".{fraction}" if fraction else "")
        try:
            parsed = Decimal(normalized)
        except InvalidOperation:
            return None
        return -parsed if negative else parsed

    @classmethod
    def _diagnostic_money_value(cls, text: str) -> Optional[Decimal]:
        """Read cents from a corrupt token only for ledger diagnostics.

        The digit fingerprint is never written back to Table_Data.  It merely
        lets a malformed token such as ``3.225.65`` trigger a direct crop OCR
        search for its printed sign and separators.
        """
        strict = cls._strict_money_value(text)
        if strict is not None:
            return strict
        digits = cls._money_digit_fingerprint(text)
        if len(digits) < 3:
            return None
        try:
            parsed = Decimal(digits) / Decimal("100")
        except InvalidOperation:
            return None
        return -parsed if cls._has_negative_money_marker(text) else parsed

    @classmethod
    def _numeric_refinement_targets(
        cls,
        columns: List[TableColumn],
        values_by_row: List[List[str]],
        previous_values: Optional[List[str]] = None,
    ) -> set[Tuple[int, int]]:
        """Find cells that warrant direct OCR, never values to infer.

        Invalid monetary syntax is direct evidence.  A running-balance mismatch
        is only a diagnostic trigger for re-reading the printed balance cell;
        arithmetic is never used as a replacement value.
        """
        if not values_by_row:
            return set()

        amount_indices = [
            idx for idx, column in enumerate(columns)
            if cls._is_amount_header(column.header)
        ]
        targets: set[Tuple[int, int]] = set()
        for row_idx, values in enumerate(values_by_row):
            for column_idx in amount_indices:
                if column_idx >= len(values):
                    continue
                value = str(values[column_idx] or "").strip()
                if cls._money_digit_fingerprint(value) and cls._strict_money_value(value) is None:
                    targets.add((row_idx, column_idx))

        targets.update(cls._ledger_balance_mismatch_targets(
            columns,
            values_by_row,
            previous_values,
        ))
        return targets

    @classmethod
    def _ledger_balance_mismatch_targets(
        cls,
        columns: List[TableColumn],
        values_by_row: List[List[str]],
        previous_values: Optional[List[str]] = None,
    ) -> set[Tuple[int, int]]:
        targets: set[Tuple[int, int]] = set()

        balance_idx = next(
            (
                idx for idx, column in enumerate(columns)
                if "balance" in column.header.casefold()
                or "saldo" in cls._header_token_base(column.header)
            ),
            None,
        )
        debit_idx = next(
            (
                idx for idx, column in enumerate(columns)
                if any(
                    token in column.header.casefold()
                    for token in ("debit", "withdraw", "money out", "soll", "cargo")
                )
            ),
            None,
        )
        credit_idx = next(
            (
                idx for idx, column in enumerate(columns)
                if any(
                    token in column.header.casefold()
                    for token in ("credit", "deposit", "money in", "haben", "abono")
                )
            ),
            None,
        )
        if balance_idx is None or (debit_idx is None and credit_idx is None):
            return targets

        prior = previous_values
        for row_idx, values in enumerate(values_by_row):
            if prior is None:
                prior = values
                continue
            if max(
                balance_idx,
                debit_idx if debit_idx is not None else 0,
                credit_idx if credit_idx is not None else 0,
            ) >= len(values) or balance_idx >= len(prior):
                prior = values
                continue

            previous_balance = cls._diagnostic_money_value(prior[balance_idx])
            current_balance = cls._diagnostic_money_value(values[balance_idx])

            def directed_amount(column_idx: Optional[int]) -> Optional[Decimal]:
                if column_idx is None:
                    return Decimal("0")
                cell = str(values[column_idx] or "").strip()
                if not cell:
                    return Decimal("0")
                parsed = cls._diagnostic_money_value(cell)
                return abs(parsed) if parsed is not None else None

            debit = directed_amount(debit_idx)
            credit = directed_amount(credit_idx)
            if None not in (previous_balance, current_balance, debit, credit):
                expected = previous_balance - debit + credit
                if abs(current_balance - expected) >= Decimal("0.005"):
                    targets.add((row_idx, balance_idx))
            prior = values

        return targets

    def _refine_ocr_numeric_cells(
        self,
        page: LayoutPage,
        columns: List[TableColumn],
        word_buckets: List[List[List[LayoutWord]]],
        values_by_row: List[List[str]],
    ) -> None:
        """Replace a suspect Table_Data token only with two direct OCR reads."""
        if (
            page.source not in {"ocr", "hybrid"}
            or not self.use_paddleocr
            or not self._pdf_path
        ):
            return

        previous_values: Optional[List[str]] = None
        if self.table_rows:
            previous_row = self.table_rows[-1]
            if len(previous_row.values) == len(columns):
                previous_values = previous_row.values
        targets = self._numeric_refinement_targets(
            columns,
            values_by_row,
            previous_values,
        )
        if not targets:
            return

        try:
            from .paddleocr_processor import PaddleOCRProcessor

            processor = PaddleOCRProcessor(use_table_structure=False)
        except Exception as exc:
            logger.warning("Cell-local numeric OCR unavailable on page %s: %s", page.page_number, exc)
            self._numeric_ocr_unresolved_count += len(targets)
            return

        image_cache: Dict[Tuple[Any, ...], Any] = {}
        unresolved: set[Tuple[int, int]] = set()
        for row_idx, column_idx in sorted(targets):
            current_targets = self._numeric_refinement_targets(
                columns,
                values_by_row,
                previous_values,
            )
            if (row_idx, column_idx) not in current_targets:
                # A preceding direct correction can remove the apparent
                # mismatch on the next row; do not OCR that now-consistent row.
                continue
            if (
                row_idx >= len(values_by_row)
                or row_idx >= len(word_buckets)
                or column_idx >= len(values_by_row[row_idx])
                or column_idx >= len(word_buckets[row_idx])
            ):
                unresolved.add((row_idx, column_idx))
                continue
            original = values_by_row[row_idx][column_idx]
            words = word_buckets[row_idx][column_idx]
            ledger_targets = self._ledger_balance_mismatch_targets(
                columns,
                values_by_row,
                previous_values,
            )
            replacement = self._agreed_numeric_crop_candidate(
                page,
                columns[column_idx],
                words,
                original,
                processor,
                image_cache,
                prefer_explicit_negative=(row_idx, column_idx) in ledger_targets,
            )
            if replacement is None or replacement == original:
                unresolved.add((row_idx, column_idx))
                continue
            values_by_row[row_idx][column_idx] = replacement
            self._numeric_ocr_refinement_count += 1

        remaining_targets = self._numeric_refinement_targets(
            columns,
            values_by_row,
            previous_values,
        )
        unresolved.update(remaining_targets)
        self._numeric_ocr_unresolved_count += len(unresolved)

    def _agreed_numeric_crop_candidate(
        self,
        page: LayoutPage,
        column: TableColumn,
        words: List[LayoutWord],
        original: str,
        processor: Any,
        image_cache: Dict[Tuple[Any, ...], Any],
        prefer_explicit_negative: bool = False,
    ) -> Optional[str]:
        if not words or not self._money_digit_fingerprint(original):
            return None

        agreed_candidates: List[str] = []
        for half_height in (30.0, 35.0, 40.0, 45.0):
            base_readings = [
                self._read_numeric_cell_crop(
                    page,
                    column,
                    words,
                    original,
                    processor,
                    image_cache,
                    dpi,
                    half_height,
                )
                for dpi in (150, 250)
            ]
            readings = [
                reading
                if reading is not None
                and self._is_safe_direct_numeric_candidate(original, reading)
                else None
                for reading in base_readings
            ]
            candidate = readings[0] if readings[0] and readings[0] == readings[1] else None
            seek_negative = (
                prefer_explicit_negative
                and not self._has_negative_money_marker(original)
            )
            if candidate is None:
                candidate_counts: Dict[str, int] = {}
                for reading in readings:
                    if reading:
                        candidate_counts[reading] = candidate_counts.get(reading, 0) + 1

                # PyMuPDF and Poppler rasterize the thin minus at slightly
                # different sub-pixel positions.  Consult additional render
                # scales only when the primary pair did not agree, and still
                # require two independent DPI readings of the exact token.
                for dpi in (180, 165, 200, 220, 300):
                    reading = self._read_numeric_cell_crop(
                        page,
                        column,
                        words,
                        original,
                        processor,
                        image_cache,
                        dpi,
                        half_height,
                    )
                    if (
                        reading is None
                        or not self._is_safe_direct_numeric_candidate(original, reading)
                    ):
                        continue
                    candidate_counts[reading] = candidate_counts.get(reading, 0) + 1
                    if candidate_counts[reading] < 2:
                        continue
                    if seek_negative and not self._has_negative_money_marker(reading):
                        continue
                    candidate = reading
                    break

            if candidate and candidate != original:
                agreed_candidates.append(candidate)
                if (
                    not prefer_explicit_negative
                    or self._has_negative_money_marker(original)
                    or self._has_negative_money_marker(candidate)
                ):
                    return candidate

        full_page_candidate = self._agreed_numeric_full_page_candidate(
            page,
            words,
            original,
            processor,
            image_cache,
        )
        if full_page_candidate is not None:
            if (
                not prefer_explicit_negative
                or self._has_negative_money_marker(original)
                or self._has_negative_money_marker(full_page_candidate)
            ):
                return full_page_candidate

        if not agreed_candidates:
            return None
        if prefer_explicit_negative and not self._has_negative_money_marker(original):
            # A positive re-read cannot resolve a diagnosed missing-sign case.
            # Leave the source value for review unless two OCR passes expose
            # the same printed negative marker.
            return None
        return agreed_candidates[0]

    def _agreed_numeric_full_page_candidate(
        self,
        page: LayoutPage,
        words: List[LayoutWord],
        original: str,
        processor: Any,
        image_cache: Dict[Tuple[Any, ...], Any],
    ) -> Optional[str]:
        """Rare fallback: require two alternate-DPI full-page OCR readings."""
        readings = [
            self._read_numeric_full_page_candidate(
                page,
                words,
                original,
                processor,
                image_cache,
                dpi,
            )
            for dpi in (180, 220)
        ]
        if readings[0] and readings[0] == readings[1] and readings[0] != original:
            return readings[0]
        return None

    def _read_numeric_full_page_candidate(
        self,
        page: LayoutPage,
        words: List[LayoutWord],
        original: str,
        processor: Any,
        image_cache: Dict[Tuple[Any, ...], Any],
        dpi: int,
    ) -> Optional[str]:
        image_key = (page.page_number, dpi)
        image = image_cache.get(image_key)
        if image is None:
            try:
                image = self._render_page_image(self._pdf_path, page.page_number, dpi)
            except Exception as exc:
                logger.warning(
                    "Unable to render page %s for numeric OCR at %s DPI: %s",
                    page.page_number,
                    dpi,
                    exc,
                )
                return None
            image_cache[image_key] = image

        result_key = ("full-page-numeric-ocr", page.page_number, dpi)
        results = image_cache.get(result_key)
        if results is None:
            try:
                results = processor.extract_with_coordinates(image, confidence_threshold=0.0)
            except Exception as exc:
                logger.warning("Full-page numeric OCR failed on page %s: %s", page.page_number, exc)
                return None
            image_cache[result_key] = results

        target_x = statistics.median((word.x0 + word.x1) / 2.0 for word in words)
        target_y = statistics.median(word.center_y for word in words)
        target_width = max(word.x1 for word in words) - min(word.x0 for word in words)
        target_height = max(word.bottom for word in words) - min(word.top for word in words)
        scale_x = page.width / max(float(image.width), 1.0)
        scale_y = page.height / max(float(image.height), 1.0)
        candidates: List[Tuple[float, float, str]] = []
        for result in results or []:
            text = re.sub(r"\s+", "", str(result.get("text") or "")).strip("|[]{}")
            bbox = result.get("bbox")
            try:
                x0, y0, x1, y1 = [float(value) for value in bbox]
                confidence = float(result.get("confidence", 0.0))
            except (TypeError, ValueError):
                continue
            if not text or confidence < 0.80:
                continue
            center_x = ((x0 + x1) / 2.0) * scale_x
            center_y = ((y0 + y1) / 2.0) * scale_y
            if abs(center_x - target_x) > max(40.0, target_width * 0.80):
                continue
            if abs(center_y - target_y) > max(12.0, target_height * 0.80):
                continue
            candidate = self._preserve_negative_money_marker(original, text)
            if not self._is_safe_direct_numeric_candidate(original, candidate):
                continue
            distance = abs(center_x - target_x) + abs(center_y - target_y)
            candidates.append((distance, -confidence, candidate))

        if not candidates:
            return None
        return min(candidates)[2]

    @classmethod
    def _is_safe_direct_numeric_candidate(cls, original: str, candidate: str) -> bool:
        if cls._strict_money_value(candidate) is None:
            return False
        if cls._money_digit_fingerprint(candidate) != cls._money_digit_fingerprint(original):
            return False
        if cls._has_negative_money_marker(original) and not cls._has_negative_money_marker(candidate):
            return False
        original_currencies = set(re.findall(r"[$£€₹]", str(original)))
        candidate_currencies = set(re.findall(r"[$£€₹]", str(candidate)))
        return not original_currencies or candidate_currencies == original_currencies

    def _read_numeric_cell_crop(
        self,
        page: LayoutPage,
        column: TableColumn,
        words: List[LayoutWord],
        original: str,
        processor: Any,
        image_cache: Dict[Tuple[Any, ...], Any],
        dpi: int,
        half_height: float,
    ) -> Optional[str]:
        cache_key = (page.page_number, dpi)
        image = image_cache.get(cache_key)
        if image is None:
            try:
                image = self._render_page_image(self._pdf_path, page.page_number, dpi)
            except Exception as exc:
                logger.warning(
                    "Unable to render page %s for numeric OCR at %s DPI: %s",
                    page.page_number,
                    dpi,
                    exc,
                )
                return None
            image_cache[cache_key] = image

        scale_x = image.width / max(page.width, 1.0)
        scale_y = image.height / max(page.height, 1.0)
        center_y = statistics.median(word.center_y for word in words)
        left = max(0.0, min(column.x0, min(word.x0 for word in words)) - 30.0)
        right = min(page.width, max(column.right, max(word.x1 for word in words)) + 30.0)
        top = max(0.0, center_y - half_height)
        bottom = min(page.height, center_y + half_height)
        crop_box = (
            int(left * scale_x),
            int(top * scale_y),
            int(right * scale_x),
            int(bottom * scale_y),
        )
        if crop_box[2] <= crop_box[0] or crop_box[3] <= crop_box[1]:
            return None
        crop = image.crop(crop_box)
        try:
            results = processor.extract_with_coordinates(crop, confidence_threshold=0.0)
        except Exception as exc:
            logger.warning("Cell-local numeric OCR failed on page %s: %s", page.page_number, exc)
            return None
        return self._numeric_candidate_from_crop_results(
            results,
            original,
            crop.width,
            crop.height,
        )

    @classmethod
    def _numeric_candidate_from_crop_results(
        cls,
        results: List[Dict[str, Any]],
        original: str,
        crop_width: int,
        crop_height: int,
    ) -> Optional[str]:
        """Select the direct OCR money token nearest the crop's row center."""
        del crop_width  # The vertical row anchor is the discriminating axis.
        items: List[Dict[str, Any]] = []
        for result in results or []:
            text = re.sub(r"\s+", "", str(result.get("text") or "")).strip("|[]{}")
            bbox = result.get("bbox")
            try:
                x0, y0, x1, y1 = [float(value) for value in bbox]
                confidence = float(result.get("confidence", 0.0))
            except (TypeError, ValueError):
                continue
            if not text or confidence < 0.80 or x1 <= x0 or y1 <= y0:
                continue
            items.append({
                "text": text,
                "x0": x0,
                "x1": x1,
                "y0": y0,
                "y1": y1,
                "confidence": confidence,
            })

        candidate_rows: List[Tuple[str, float, float]] = []
        for item in items:
            candidate_rows.append((
                str(item["text"]),
                (float(item["y0"]) + float(item["y1"])) / 2.0,
                float(item["confidence"]),
            ))

        line_groups: List[List[Dict[str, Any]]] = []
        for item in sorted(items, key=lambda value: ((value["y0"] + value["y1"]) / 2.0, value["x0"])):
            center_y = (float(item["y0"]) + float(item["y1"])) / 2.0
            for group in line_groups:
                group_center = statistics.median(
                    (float(member["y0"]) + float(member["y1"])) / 2.0
                    for member in group
                )
                group_height = statistics.median(
                    float(member["y1"]) - float(member["y0"])
                    for member in group
                )
                if abs(center_y - group_center) <= max(5.0, group_height * 0.55):
                    group.append(item)
                    break
            else:
                line_groups.append([item])

        for group in line_groups:
            if len(group) < 2:
                continue
            ordered = sorted(group, key=lambda value: value["x0"])
            candidate_rows.append((
                "".join(str(item["text"]) for item in ordered),
                statistics.median(
                    (float(item["y0"]) + float(item["y1"])) / 2.0
                    for item in ordered
                ),
                min(float(item["confidence"]) for item in ordered),
            ))

        fingerprint = cls._money_digit_fingerprint(original)
        original_negative = cls._has_negative_money_marker(original)
        original_currencies = set(re.findall(r"[$£€₹]", str(original)))
        eligible: List[Tuple[float, float, str]] = []
        for candidate, center_y, confidence in candidate_rows:
            if cls._money_digit_fingerprint(candidate) != fingerprint:
                continue
            if cls._strict_money_value(candidate) is None:
                continue
            if original_negative and not cls._has_negative_money_marker(candidate):
                # The source layer already contains the thin negative mark.
                # Preserve it while the crop supplies only punctuation; this
                # never invents or removes a sign.
                candidate = cls._preserve_negative_money_marker(original, candidate)
            candidate_currencies = set(re.findall(r"[$£€₹]", candidate))
            if original_currencies and candidate_currencies != original_currencies:
                continue
            distance = abs(center_y - (crop_height / 2.0))
            if distance > crop_height * 0.30:
                continue
            eligible.append((distance, -confidence, candidate))

        if not eligible:
            return None
        return min(eligible)[2]

    @staticmethod
    def _table_schema_signature(columns: List[TableColumn]) -> Tuple[str, ...]:
        """Return a stable, immutable identity for the extraction-time schema."""
        return tuple(
            re.sub(r"\s+", " ", column.header.strip()).casefold()
            for column in columns
        )

    def _detect_table_header_line(self, page: LayoutPage) -> Optional[LayoutLine]:
        best_line: Optional[LayoutLine] = None
        best_score = 0.0
        lines_by_index = {line.index: line for line in page.lines}

        for source_line in page.lines:
            line = self._merge_ocr_header_fragments(page, source_line)
            if len(line.words) < 3 or self._is_separator_line(line):
                continue

            words = [word.text.strip() for word in line.words if word.text.strip()]
            if not words:
                continue

            alpha_words = sum(1 for word in words if re.search(r"[A-Za-z_]", word))
            numeric_words = sum(1 for word in words if re.search(r"\d", word))
            date_like_words = sum(1 for word in words if self._looks_like_date_value(word))
            amount_like_words = sum(1 for word in words if self._looks_like_amount_value(word))
            if alpha_words < 2 or numeric_words >= alpha_words:
                continue
            if date_like_words or amount_like_words >= 2:
                continue

            previous_separator = any(
                self._is_separator_line(lines_by_index[idx])
                for idx in range(max(1, line.index - 2), line.index)
                if idx in lines_by_index
            )
            next_separator = any(
                self._is_separator_line(lines_by_index[idx])
                for idx in range(line.index + 1, line.index + 3)
                if idx in lines_by_index
            )
            word_span = max(word.x1 for word in line.words) - min(word.x0 for word in line.words)
            uppercaseish = sum(
                1 for word in words
                if not re.search(r"[a-z]", word) or "_" in word
            )
            # Known column vocabulary outranks e.g. an ALL-CAPS bank title —
            # a short "Date  Description  Amount" line is a header, the title
            # above it is not.
            vocab_hits = self._header_vocabulary_hits(line)

            score = len(words) + alpha_words + (word_span / 120.0) + (uppercaseish / 2.0) + (3.0 * vocab_hits)
            if previous_separator:
                score += 4.0
            if next_separator:
                score += 5.0

            if score > best_score:
                best_score = score
                best_line = line

        if best_score < 8.0 or best_line is None:
            return None
        return self._merge_ocr_header_fragments(page, best_line)

    def _merge_ocr_header_fragments(
        self,
        page: LayoutPage,
        header_line: LayoutLine,
    ) -> LayoutLine:
        """Join a header split into overlapping OCR lines.

        A perspective-skewed scan can put the left side of a single visual
        header (for example ``Date / Description``) on the next OCR line while
        its amount labels remain on the first.  Only adjacent, vertically
        overlapping, lexical header fragments with disjoint word boxes are
        joined.  The page's original lines remain untouched, so ``Exact_Copy``
        still reflects the OCR layer exactly as it was extracted.
        """
        if page.source not in {"ocr", "hybrid"}:
            return header_line

        base_hits = self._header_vocabulary_hits(header_line)
        if base_hits < 1:
            return header_line

        fragments: List[LayoutLine] = []
        for candidate in page.lines:
            if candidate.index == header_line.index:
                continue
            if abs(candidate.index - header_line.index) > 1:
                continue
            candidate_hits = self._header_vocabulary_hits(candidate)
            if candidate_hits <= 0:
                continue
            if any(
                self._looks_like_date_value(word.text)
                or self._looks_like_amount_value(word.text)
                for word in candidate.words
            ):
                continue

            overlap = min(header_line.bottom, candidate.bottom) - max(
                header_line.top,
                candidate.top,
            )
            overlap_ratio = overlap / max(
                1.0,
                min(header_line.height, candidate.height),
            )
            if overlap_ratio < 0.15:
                continue

            boxes_collide = any(
                min(base_word.x1, fragment_word.x1)
                - max(base_word.x0, fragment_word.x0)
                > 2.0
                for base_word in header_line.words
                for fragment_word in candidate.words
            )
            if boxes_collide:
                continue
            fragments.append(candidate)

        if not fragments:
            return header_line
        if base_hits + sum(self._header_vocabulary_hits(line) for line in fragments) < 3:
            return header_line

        words = sorted(
            header_line.words + [word for line in fragments for word in line.words],
            key=lambda word: word.x0,
        )
        return LayoutLine(
            page=page.page_number,
            index=header_line.index,
            top=min(word.top for word in words),
            bottom=max(word.bottom for word in words),
            center_y=sum(word.center_y for word in words) / len(words),
            words=words,
            text=self._line_text(words),
        )

    @staticmethod
    def _header_vocabulary_hits(line: LayoutLine) -> int:
        header_vocabulary = {
            "date", "description", "amount", "balance", "debit", "debits",
            "credit", "credits", "withdrawals", "withdrawal", "deposits", "deposit",
            "reference", "particulars", "narration", "details", "value", "money",
            "out", "in", "type", "cheque", "chq", "txn", "datum", "omschrijving",
            "af", "bij", "saldo", "buchungstag", "verwendungszweck", "soll", "haben",
            "betrag",
        }
        return sum(
            1
            for word in line.words
            if LayoutReplicaParser._header_token_base(word.text) in header_vocabulary
        )

    @staticmethod
    def _header_token_base(text: str) -> str:
        normalized = text.casefold().strip()
        normalized = re.sub(r"\(\s*[-+]\s*\)$", "", normalized).strip()
        return normalized.strip(".,:;")

    def _columns_from_transaction_lines(self, page: LayoutPage) -> List[TableColumn]:
        candidates = [
            line for line in page.lines
            if len(line.words) >= 4 and self._line_starts_with_date(line)
        ]
        if not candidates:
            return []

        candidate = max(
            candidates,
            key=lambda line: (
                len(line.words),
                max(word.x1 for word in line.words) - min(word.x0 for word in line.words),
            ),
        )
        words = sorted(candidate.words, key=lambda item: item.x0)
        if len(words) < 4:
            return []

        columns: List[TableColumn] = []
        for idx, word in enumerate(words):
            if idx == 0:
                left = max(0.0, word.x0 - 6.0)
            else:
                previous = words[idx - 1]
                left = (previous.x1 + word.x0) / 2.0

            if idx == len(words) - 1:
                right = min(page.width, word.x1 + 18.0)
            else:
                nxt = words[idx + 1]
                if idx == 1:
                    right = max(word.x1, nxt.x0 - 2.0)
                else:
                    right = (word.x1 + nxt.x0) / 2.0

            columns.append(TableColumn(
                header=f"Column {idx + 1}",
                x0=word.x0,
                x1=word.x1,
                left=left,
                right=right,
            ))
        return columns

    def _prepend_missing_date_column(
        self,
        page: LayoutPage,
        columns: List[TableColumn],
    ) -> List[TableColumn]:
        """Recover a Date header omitted by OCR when row geometry proves it exists."""
        if not columns or any("date" in column.header.casefold() for column in columns):
            return columns
        first_column = columns[0]
        if not self._is_wide_text_header(first_column.header):
            return columns
        date_words = [
            word
            for word in page.words
            if self._looks_like_date_value(word.text) and word.x1 < first_column.x0
        ]
        if len(date_words) < 2:
            return columns
        date_x0 = statistics.median(word.x0 for word in date_words)
        date_x1 = statistics.median(word.x1 for word in date_words)
        boundary = (date_x1 + first_column.x0) / 2.0
        if boundary <= date_x0 or boundary >= first_column.x0:
            return columns
        first_column.left = boundary
        return [
            TableColumn(
                header="Date",
                x0=date_x0,
                x1=date_x1,
                left=max(0.0, min(word.x0 for word in date_words) - 8.0),
                right=boundary,
            ),
            *columns,
        ]

    def _line_starts_with_date(self, line: LayoutLine) -> bool:
        if not line.words:
            return False
        first_word = sorted(line.words, key=lambda item: item.x0)[0].text
        return self._looks_like_date_value(first_word)

    @staticmethod
    def _looks_like_date_value(text: str) -> bool:
        # OCR often keeps a table border or terminal full stop attached to an
        # otherwise valid date token.
        value = text.strip().strip("|[](){} ").rstrip(".,;:")
        if not value:
            return False
        # Numeric dates: 06/03/2026, 03-06-2026, 03.06.2026 (3 parts, any of -/. )
        if re.fullmatch(r"\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}", value):
            return True
        # Short numeric dates without a year (US card style: 06/11). Slash or
        # dash only — a dot here would swallow decimals like "1.5".
        if re.fullmatch(r"\d{1,2}[-/]\d{1,2}", value):
            return True
        # Month-name dates: "3 Jun 2026", "30 June 2026", "Jun 3, 2026", "3 Jun"
        month = r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?"
        if re.fullmatch(rf"\d{{1,2}}\s+{month}(?:\s+\d{{2,4}})?", value, re.IGNORECASE):
            return True
        if re.fullmatch(rf"{month}\s+\d{{1,2}},?(?:\s+\d{{2,4}})?", value, re.IGNORECASE):
            return True
        return False

    @staticmethod
    def _looks_like_amount_value(text: str) -> bool:
        value = text.strip().strip("|[]{} ")
        if value.startswith("(") and value.endswith(")"):
            value = value[1:-1].strip()
        value = value.lstrip("+-").strip()
        if value[:1] in ("$", "£", "€"):
            value = value[1:].strip()
        if not value:
            return False
        # OCR can retain sentence/table punctuation after a complete monetary
        # token (for example ``-5,711.97.``).  Remove one terminal mark only
        # when a decimal-looking suffix already precedes it.
        if re.search(r"[.,]\d{1,4}[.,;:]$", value):
            value = value[:-1]
        # 2,480.00 · 1.234,56 · 2480 · 0.84 — thousands groups with , or .
        # and an optional 1-4 digit decimal part with the other separator.
        return bool(re.fullmatch(r"(?:\d{1,3}(?:[.,]\d{3})+|\d+)(?:[.,]\d{1,4})?", value))

    @staticmethod
    def _is_separator_line(line: LayoutLine) -> bool:
        text = "".join(word.text for word in line.words).strip()
        if len(text) < 12:
            return False
        separator_chars = sum(1 for ch in text if ch in "-_=—–.")
        return separator_chars / max(len(text), 1) >= 0.85

    def _columns_from_header_line(self, header_line: LayoutLine, page_width: float) -> List[TableColumn]:
        groups: List[List[LayoutWord]] = []
        current: List[LayoutWord] = []
        char_widths = [
            word.width / max(len(word.text), 1)
            for word in header_line.words
            if word.text
        ]
        median_char_width = statistics.median(char_widths) if char_widths else 5.0
        merge_gap = max(5.5, min(8.5, median_char_width * 1.45))

        for word in sorted(header_line.words, key=lambda item: item.x0):
            if not current:
                current = [word]
                continue
            gap = word.x0 - current[-1].x1
            qualifier = word.text.strip().replace(" ", "") in {"(-)", "(+)"}
            current_label = " ".join(item.text for item in current).casefold()
            if (
                qualifier
                and any(label in current_label for label in ("debit", "credit"))
                and gap <= max(24.0, merge_gap)
            ):
                current.append(word)
            elif gap <= merge_gap:
                current.append(word)
            else:
                groups.append(current)
                current = [word]
        if current:
            groups.append(current)

        header_cells = []
        for group in groups:
            text = " ".join(word.text.strip() for word in group if word.text.strip())
            if not text:
                continue
            header_cells.append({
                "header": self._normalize_header_label(text),
                "x0": min(word.x0 for word in group),
                "x1": max(word.x1 for word in group),
            })

        columns: List[TableColumn] = []
        for idx, cell in enumerate(header_cells):
            if idx == 0:
                left = max(0.0, cell["x0"] - 8.0)
            else:
                prev = header_cells[idx - 1]
                left = (prev["x1"] + cell["x0"]) / 2.0

            if idx == len(header_cells) - 1:
                right = min(page_width, cell["x1"] + 20.0)
            else:
                nxt = header_cells[idx + 1]
                if self._is_wide_text_header(str(cell["header"])):
                    right = max(cell["x1"], nxt["x0"] - 2.0)
                else:
                    right = (cell["x1"] + nxt["x0"]) / 2.0

            columns.append(TableColumn(
                header=str(cell["header"]),
                x0=float(cell["x0"]),
                x1=float(cell["x1"]),
                left=float(left),
                right=float(right),
            ))
        return columns

    @staticmethod
    def _normalize_header_label(header: str) -> str:
        """Canonicalize harmless OCR punctuation around amount qualifiers."""
        normalized = re.sub(r"\s+", " ", header.strip())
        match = re.fullmatch(
            r"(debits?|credits?)\s*[.,:;]?\s*\(\s*([-+])\s*\)",
            normalized,
            re.IGNORECASE,
        )
        if not match:
            return normalized
        label = match.group(1)
        return f"{label} ({match.group(2)})"

    @staticmethod
    def _is_wide_text_header(header: str) -> bool:
        normalized = header.lower()
        tokens = ("description", "details", "narration", "particular", "remarks", "reference", "ref")
        return any(token in normalized for token in tokens)

    def _assign_line_to_table_columns(self, line: LayoutLine, columns: List[TableColumn]) -> List[str]:
        buckets: List[List[LayoutWord]] = [[] for _ in columns]
        for word in sorted(line.words, key=lambda item: item.x0):
            col_idx = self._table_column_index_for_word(word, columns)
            if col_idx is not None:
                buckets[col_idx].append(word)

        self._move_narrow_column_overflow(buckets, columns)
        return [self._join_words(words) for words in buckets]

    def _is_table_relevant_row(
        self,
        line: LayoutLine,
        values: List[str],
        columns: List[TableColumn],
        saw_data_row: bool,
    ) -> bool:
        text = line.text.strip()
        if self._is_document_context_line(text):
            return False

        populated_indices = [idx for idx, value in enumerate(values) if value]
        if not populated_indices:
            return False

        first_two_values = values[:2]
        if any(self._looks_like_date_value(value) for value in first_two_values if value):
            return True

        amount_col_indices = [
            idx for idx, column in enumerate(columns)
            if self._is_amount_header(column.header)
        ]
        if any(
            idx in amount_col_indices and self._contains_amount_value(values[idx])
            for idx in populated_indices
        ):
            return True

        if len(populated_indices) == 1 and saw_data_row:
            column = columns[populated_indices[0]]
            return self._is_wide_text_header(column.header) or populated_indices[0] in {1, 2}

        return False

    @staticmethod
    def _is_document_context_line(text: str) -> bool:
        normalized = re.sub(r"\s+", " ", text.strip().lower())
        if not normalized:
            return True
        context_patterns = (
            r"^page\s*:",
            r"statement of account",
            r"account statement",
            r"cardmember statement",
            r"account number",
            r"period from",
            r"period to",
            r"nominee\s*:",
            r"^branch\s*:",
            r"\bindian rupees\b",
            r"\bbank ltd\b",
            r"\bbank limited\b",
            r"closing balance includes",
            r"\bgstn\b",
            r"\bgstin\b",
            r"https?://",
            r"registered office",
            r"computer generated",
            r"system generated",
        )
        return any(re.search(pattern, normalized) for pattern in context_patterns)

    @staticmethod
    def _is_amount_header(header: str) -> bool:
        normalized = header.lower()
        if any(token in normalized for token in ("debit", "credit", "balance", "amount", "withdraw", "deposit")):
            return True
        # UK statements split amounts into "Money out" / "Money in".
        if "money out" in normalized or "money in" in normalized:
            return True
        # Short EU headers (NL: Af/Bij/Saldo, DE: Soll/Haben/Betrag) need
        # whole-word matches — substring checks would false-positive.
        words = set(re.split(r"[^a-zäöüß]+", normalized))
        return bool(words & {"af", "bij", "saldo", "soll", "haben", "betrag", "importe", "cargo", "abono"})

    def _contains_amount_value(self, text: str) -> bool:
        return any(self._looks_like_amount_value(part) for part in str(text).split())

    @staticmethod
    def _table_column_index_for_word(word: LayoutWord, columns: List[TableColumn]) -> Optional[int]:
        center = (word.x0 + word.x1) / 2.0
        best_idx = None
        best_key = None
        for idx, column in enumerate(columns):
            overlap = min(word.x1, column.right) - max(word.x0, column.left)
            if overlap <= 0:
                continue
            # Snap to the column holding most of the word's width; on exact
            # ties keep the column containing the word's center.
            key = (overlap, 1 if column.left <= center < column.right else 0)
            if best_key is None or key > best_key:
                best_idx, best_key = idx, key
        if best_idx is not None:
            return best_idx
        for idx, column in enumerate(columns):
            if column.left <= center < column.right:
                return idx
        if columns and center >= columns[-1].right and word.x0 <= columns[-1].right + 16.0:
            return len(columns) - 1
        return None

    @staticmethod
    def _move_narrow_column_overflow(buckets: List[List[LayoutWord]], columns: List[TableColumn]) -> None:
        for idx, column in enumerate(columns[:-1]):
            if column.width >= 45.0 or len(buckets[idx]) <= 1:
                continue
            keep = buckets[idx][:1]
            overflow = buckets[idx][1:]
            buckets[idx] = keep
            buckets[idx + 1] = sorted(overflow + buckets[idx + 1], key=lambda item: item.x0)

    @staticmethod
    def _join_words(words: List[LayoutWord]) -> str:
        return " ".join(word.text for word in sorted(words, key=lambda item: item.x0)).strip()

    def _is_continuation_row(
        self,
        line: LayoutLine,
        values: List[str],
        columns: List[TableColumn],
        page_rows: List[TableReplicaRow],
    ) -> bool:
        """A wrapped fragment of the previous row: no date, no amounts, and
        every populated cell sits in a text-ish column."""
        if not page_rows:
            return False
        if self._is_document_context_line(line.text.strip()):
            return False
        populated = [idx for idx, value in enumerate(values) if value]
        if not populated:
            return False
        if any(self._looks_like_date_value(value) for value in values[:2] if value):
            return False
        if "date" in columns[0].header.casefold():
            # In an explicit Date schema, OCR commonly emits the description
            # and amounts from one visual transaction on separate lines. Once
            # a dated row exists, an undated fragment belongs to that row.
            return True
        for idx in populated:
            if self._is_amount_header(columns[idx].header) and self._contains_amount_value(values[idx]):
                return False
            if not self._is_text_column(columns[idx], idx):
                return False
        return True

    def _is_text_column(self, column: TableColumn, idx: int) -> bool:
        header = column.header.lower()
        if self._is_wide_text_header(column.header):
            return True
        if any(token in header for token in ("reference", "ref", "remarks")):
            return True
        if header.startswith("column"):
            # positional layouts: description/reference usually sit at index 1-2
            return idx in {1, 2}
        return False

    @staticmethod
    def _merge_continuation_row(parent: TableReplicaRow, values: List[str]) -> None:
        for idx, value in enumerate(values):
            if not value or idx >= len(parent.values):
                continue
            parent.values[idx] = f"{parent.values[idx]} {value}".strip()

    def _write_table_replica_sheet(self, wb: Workbook) -> None:
        groups: List[Tuple[List[str], List[TableReplicaRow]]] = []
        group_indices: Dict[Tuple[str, ...], int] = {}

        for row in self.table_rows:
            headers = list(row.schema_headers)
            if not headers:
                headers = [column.header for column in self.table_columns]
            if len(headers) < len(row.values):
                headers.extend(
                    f"Column {idx}"
                    for idx in range(len(headers) + 1, len(row.values) + 1)
                )

            signature = row.schema_signature or tuple(
                re.sub(r"\s+", " ", header.strip()).casefold()
                for header in headers
            )
            if signature not in group_indices:
                group_indices[signature] = len(groups)
                groups.append((headers, []))
            groups[group_indices[signature]][1].append(row)

        if not groups:
            headers = [column.header for column in self.table_columns]
            groups.append((headers or ["No table detected"], []))

        for group_number, (headers, rows) in enumerate(groups, 1):
            title = "Table_Data" if group_number == 1 else f"Table_Data_{group_number}"
            sheet = wb.create_sheet(title)

            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx)
                self._set_text_cell(cell, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="EAF2F8")
                sheet.column_dimensions[get_column_letter(col_idx)].width = self._table_column_width(header)

            for row_idx, row in enumerate(rows, 2):
                values = list(row.values)
                if len(values) < len(headers):
                    values.extend([""] * (len(headers) - len(values)))
                for col_idx, value in enumerate(values, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    self._set_text_cell(cell, value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _table_column_width(header: str) -> float:
        normalized = header.lower()
        if any(token in normalized for token in ("description", "details", "narration", "particular")):
            return 42
        if any(token in normalized for token in ("reference", "ref", "remarks")):
            return 24
        if any(token in normalized for token in ("balance", "debit", "credit", "amount")):
            return 16
        if any(token in normalized for token in ("date", "dt")):
            return 13
        return max(10, min(22, len(header) + 4))

    def _write_table_index_sheet(self, wb: Workbook) -> None:
        sheet = wb.create_sheet("Table_Index")
        headers = ["Page", "Header Line", "Columns", "Rows", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            self._set_text_cell(cell, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2F8")

        for row_idx, summary in enumerate(self.table_page_summaries, 2):
            values = [
                str(summary.get("page", "")),
                str(summary.get("header_line", "")),
                str(summary.get("columns", "")),
                str(summary.get("rows", "")),
                str(summary.get("status", "")),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                self._set_text_cell(cell, value)

        widths = [10, 14, 12, 10, 18]
        for col_idx, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_lines_sheet(self, wb: Workbook) -> None:
        sheet = wb.create_sheet("Full_Text")
        headers = ["Page", "Line", "Source", "Text"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            self._set_text_cell(cell, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2F8")
        row_idx = 2
        for page in self.pages:
            for line in page.lines:
                self._set_text_cell(sheet.cell(row=row_idx, column=1), page.page_number)
                self._set_text_cell(sheet.cell(row=row_idx, column=2), line.index)
                self._set_text_cell(sheet.cell(row=row_idx, column=3), page.source)
                self._set_text_cell(sheet.cell(row=row_idx, column=4), line.text)
                row_idx += 1
        sheet.column_dimensions["A"].width = 8
        sheet.column_dimensions["B"].width = 8
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 120

    def _write_page_index_sheet(self, wb: Workbook) -> None:
        sheet = wb.create_sheet("Page_Index")
        headers = ["Page", "Source", "Visual Lines", "Words", "Status"]
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            self._set_text_cell(cell, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2F8")

        for row_idx, page in enumerate(self.pages, 2):
            status = "ok" if page.words else "empty"
            values = [
                str(page.page_number),
                page.source,
                str(len(page.lines)),
                str(len(page.words)),
                status,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                self._set_text_cell(cell, value)

        widths = [10, 14, 14, 12, 12]
        for col_idx, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_metadata_sheet(self, wb: Workbook) -> None:
        sheet = wb.create_sheet("Extraction_Metadata")
        rows = [
            ("Mode", "layout_replica"),
            ("Source file", self.source_filename),
            ("Pages", str(len(self.pages))),
            ("Table rows", str(len(self.table_rows))),
            ("Table columns", str(len(self.table_columns))),
            ("Visual lines", str(self.extraction_metadata.row_count)),
            ("Approx. layout columns", str(self.extraction_metadata.col_count)),
            ("PDF type", self.extraction_metadata.pdf_type),
            ("Message", self.extraction_metadata.message),
        ]
        for row_idx, (key, value) in enumerate(rows, 1):
            key_cell = sheet.cell(row=row_idx, column=1)
            self._set_text_cell(key_cell, key)
            key_cell.font = Font(bold=True)
            self._set_text_cell(sheet.cell(row=row_idx, column=2), value)
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 80

    def _build_raw_table(self) -> None:
        rows = []
        for page in self.pages:
            for line in page.lines:
                rows.append([str(page.page_number), str(line.index), line.text])
        self.raw_table = {
            "columns": ["Page", "Line", "Text"],
            "rows": rows,
        }

    def _build_metadata(self) -> None:
        line_count = sum(len(page.lines) for page in self.pages)
        word_count = sum(len(page.words) for page in self.pages)
        text_pages = sum(1 for page in self.pages if page.source in {"pdf-text", "hybrid"})
        ocr_pages = sum(1 for page in self.pages if page.source in {"ocr", "hybrid"})
        empty_pages = sum(1 for page in self.pages if not page.words)
        schema_count = len({
            row.schema_signature
            for row in self.table_rows
            if row.schema_signature
        })
        max_cols = 0
        for page in self.pages:
            page_cols = int(round(page.width / self.points_per_column)) + 2
            max_cols = max(max_cols, min(self.max_columns, page_cols))

        if word_count == 0:
            confidence = "empty"
            message = "No visible text was extracted from the PDF."
        elif empty_pages:
            confidence = "low"
            message = (
                f"Exact_Copy preserved text from {len(self.pages) - empty_pages} of "
                f"{len(self.pages)} pages; review the {empty_pages} empty page(s)."
            )
        elif not self.table_rows or not self.table_columns:
            confidence = "low"
            message = (
                f"Exact_Copy preserved all {line_count} extracted source rows, but no "
                "stable table columns were detected. Review the workbook against the PDF."
            )
        elif ocr_pages:
            confidence = "low"
            message = (
                f"Exact_Copy preserved all {line_count} extracted source rows. "
                f"{ocr_pages} page(s) used OCR, so verify text and column placement."
            )
        elif schema_count > 1:
            confidence = "low"
            message = (
                f"Exact_Copy preserved all {line_count} extracted source rows. "
                f"{schema_count} table layouts were placed in separate Table_Data sheets; "
                "review their column alignment."
            )
        else:
            confidence = "good"
            message = (
                f"Exact_Copy preserved all {line_count} extracted source rows and one "
                "consistent table layout was detected."
            )

        if ocr_pages and text_pages:
            pdf_type = "hybrid"
        elif ocr_pages:
            pdf_type = "image"
        elif text_pages:
            pdf_type = "text"
        else:
            pdf_type = "unknown"

        self.extraction_metadata = LayoutExtractionMetadata(
            row_count=line_count,
            col_count=max_cols,
            has_data=word_count > 0,
            pdf_type=pdf_type,
            confidence=confidence,
            message=message,
        )

    def _build_quality_report(self) -> Dict[str, Any]:
        line_count = sum(len(page.lines) for page in self.pages)
        word_count = sum(len(page.words) for page in self.pages)
        schema_signatures = {
            row.schema_signature
            for row in self.table_rows
            if row.schema_signature
        }
        table_column_count = max(
            [len(row.schema_headers) for row in self.table_rows if row.schema_headers]
            + [len(self.table_columns)]
        )
        ocr_confidences = [
            word.confidence
            for page in self.pages
            for word in page.words
            if word.confidence is not None
        ]
        avg_ocr_conf = (
            round(sum(ocr_confidences) / len(ocr_confidences), 3)
            if ocr_confidences else 0.0
        )
        return {
            "is_proxy": True,
            "mode": "layout_replica",
            "pdf_type": self.extraction_metadata.pdf_type,
            "total_pages": len(self.pages),
            "row_count": self.extraction_metadata.row_count,
            "source_line_count": line_count,
            "source_word_count": word_count,
            "exact_copy_line_count": line_count,
            "exact_copy_word_count": word_count,
            "source_coverage_pct": 100.0 if word_count else 0.0,
            "table_row_count": len(self.table_rows),
            "table_column_count": table_column_count,
            "table_schema_count": len(schema_signatures),
            "table_page_count": sum(1 for summary in self.table_page_summaries if summary.get("rows", 0)),
            "word_count": word_count,
            "ocr_page_count": sum(1 for page in self.pages if page.source in {"ocr", "hybrid"}),
            "text_page_count": sum(1 for page in self.pages if page.source in {"pdf-text", "hybrid"}),
            "ocr_confidence_avg": avg_ocr_conf,
            "numeric_ocr_refinement_count": self._numeric_ocr_refinement_count,
            "numeric_ocr_unresolved_count": self._numeric_ocr_unresolved_count,
            "review_required": self.extraction_metadata.confidence != "good",
            # No truth PDF/XLSX comparison occurs during a live conversion, so
            # reporting a synthetic "accuracy" percentage would be misleading.
            "accuracy_proxy_pct": 0.0,
        }

    @staticmethod
    def _safe_sheet_title(title: str) -> str:
        cleaned = "".join("_" if ch in "[]:*?/\\'" else ch for ch in title)
        return cleaned[:31] or "Sheet"


def create_layout_replica_parser(
    progress_callback=None,
    quality: str = "standard",
    use_ocr: bool = True,
    use_paddleocr: bool = True,
):
    dpi = 200 if quality == "high" else 150
    return LayoutReplicaParser(
        progress_callback=progress_callback,
        dpi=dpi,
        use_ocr=use_ocr,
        use_paddleocr=use_paddleocr,
    )
