#!/usr/bin/env python3
"""Validate a generated bank-statement workbook against ledger rules.

This tool is intended for local/private benchmark PDFs. It does not need truth
CSV labels; it uses exact balance continuity and optional statement summary
figures from the source PDF.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
import sys
from typing import Any, Dict, Iterable, List, Optional


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from parsers.ledger_validation import (  # noqa: E402
    StatementSummary,
    ledger_rows_from_transactions,
    validate_ledger_rows,
)
from parsers.statement_summary import merge_summaries, parse_statement_summary_text  # noqa: E402


def load_transactions_from_workbook(path: Path) -> List[Dict[str, Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = (
        workbook["Normalized_Transactions"]
        if "Normalized_Transactions" in workbook.sheetnames
        else workbook[workbook.sheetnames[0]]
    )
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for row_index in range(2, sheet.max_row + 1):
        row = {
            str(headers[col - 1]): sheet.cell(row_index, col).value
            for col in range(1, sheet.max_column + 1)
            if headers[col - 1]
        }
        if any(value not in (None, "") for value in row.values()):
            row["_workbook_row"] = row_index
            rows.append(row)
    workbook.close()
    return rows


def extract_pdf_text_native(pdf_path: Path, page_numbers: Iterable[int]) -> str:
    import pdfplumber

    chunks: List[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_number in page_numbers:
            if 1 <= page_number <= len(pdf.pages):
                text = pdf.pages[page_number - 1].extract_text() or ""
                if text.strip():
                    chunks.append(text)
    return "\n".join(chunks)


def extract_pdf_text_ocr(pdf_path: Path, page_numbers: Iterable[int], dpi: int = 180) -> str:
    from pdf2image import convert_from_path
    import pytesseract

    chunks: List[str] = []
    for page_number in page_numbers:
        images = convert_from_path(
            str(pdf_path),
            dpi=dpi,
            first_page=page_number,
            last_page=page_number,
        )
        if images:
            chunks.append(pytesseract.image_to_string(images[0], config="--psm 6"))
    return "\n".join(chunks)


def page_sample(pdf_path: Path, sample_pages: int) -> List[int]:
    import pdfplumber

    with pdfplumber.open(str(pdf_path)) as pdf:
        total_pages = len(pdf.pages)
    pages = set(range(1, min(sample_pages, total_pages) + 1))
    pages.update(range(max(1, total_pages - sample_pages + 1), total_pages + 1))
    return sorted(pages)


def extract_statement_summary(pdf_path: Optional[Path], sample_pages: int) -> StatementSummary:
    if not pdf_path:
        return StatementSummary()

    pages = page_sample(pdf_path, sample_pages)
    native_text = extract_pdf_text_native(pdf_path, pages)
    summaries = [parse_statement_summary_text(native_text)] if native_text.strip() else []

    try:
        ocr_text = extract_pdf_text_ocr(pdf_path, pages)
        if ocr_text.strip():
            summaries.append(parse_statement_summary_text(ocr_text))
    except Exception as exc:
        print(f"Warning: OCR summary extraction failed: {exc}", file=sys.stderr)

    return merge_summaries(summaries)


def summary_to_dict(summary: StatementSummary) -> Dict[str, Any]:
    from parsers.ledger_validation import format_minor

    return {
        "opening_balance": format_minor(summary.opening_balance_minor),
        "closing_balance": format_minor(summary.closing_balance_minor),
        "total_debit": format_minor(summary.total_debit_minor),
        "total_credit": format_minor(summary.total_credit_minor),
        "debit_count": summary.debit_count,
        "credit_count": summary.credit_count,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate generated statement workbook")
    parser.add_argument("workbook", type=Path, help="Generated .xlsx workbook")
    parser.add_argument("--pdf", type=Path, help="Source PDF for summary extraction")
    parser.add_argument("--sample-pages", type=int, default=2, help="Pages to sample from start and end of PDF")
    parser.add_argument("--issue-samples", type=int, default=20, help="Maximum issues to print")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON")
    args = parser.parse_args()

    transactions = load_transactions_from_workbook(args.workbook)
    summary = extract_statement_summary(args.pdf, args.sample_pages) if args.pdf else StatementSummary()
    report = validate_ledger_rows(ledger_rows_from_transactions(transactions), summary)

    payload = {
        "workbook": str(args.workbook),
        "pdf": str(args.pdf) if args.pdf else None,
        "statement_summary": summary_to_dict(summary),
        "validation": report.to_dict(),
    }

    if args.json:
        print(json.dumps(payload, indent=2))
    else:
        print(f"Workbook: {args.workbook}")
        if args.pdf:
            print(f"PDF: {args.pdf}")
        print("Statement summary:")
        for key, value in payload["statement_summary"].items():
            print(f"  {key}: {value}")
        validation = payload["validation"]
        print("Validation:")
        print(f"  rows: {validation['row_count']}")
        print(f"  balance checks: {validation['balance_checks_passed']}/{validation['balance_checks']}")
        print(f"  balance consistency: {validation['balance_consistency_pct']}%")
        print(f"  debits: count={validation['debit_count']} total={validation['total_debit']}")
        print(f"  credits: count={validation['credit_count']} total={validation['total_credit']}")
        print(f"  valid: {validation['is_valid']}")
        issues = validation["issues"]
        if issues:
            print(f"Issues ({len(issues)} total, showing {min(len(issues), args.issue_samples)}):")
            for issue in issues[: args.issue_samples]:
                row = f" row={issue['row_index']}" if issue.get("row_index") else ""
                expected = f" expected={issue['expected']}" if issue.get("expected") else ""
                actual = f" actual={issue['actual']}" if issue.get("actual") else ""
                print(f"  - {issue['code']}{row}:{expected}{actual} {issue['message']}")

    return 0 if report.is_valid else 1


if __name__ == "__main__":
    raise SystemExit(main())
